import base64
from email import message_from_bytes
from airflow import DAG
from airflow.operators.python import PythonOperator, BranchPythonOperator
from airflow.operators.trigger_dagrun import TriggerDagRunOperator
from airflow.decorators import task
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from airflow.models import Variable
from datetime import datetime, timedelta, timezone
import os
import json
import time
import logging
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
import re
from ollama import Client
import requests
import uuid
import openpyxl
import pandas as pd
from email.mime.base import MIMEBase
from email import encoders
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import pytz
from dateutil import parser

# Configure logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
def send_fallback_email_on_failure(context):
    """
    Sends a polite "we're having technical issues" email to the end-user
    when any task in the main flow fails (especially AI response generation).
    
    UPDATED: Track retry attempts and only allow 3 retries with 20-min intervals
    """
    dag_run = context.get('dag_run')
    task_instance = context.get('task_instance')
    
    dag_id = dag_run.dag_id
    run_id = dag_run.run_id
    task_id = task_instance.task_id
    execution_date = context.get('execution_date')
    
    # ============================================================
    # STEP 1: Determine if this is a retry or original failure
    # ============================================================
    
    dag_run_conf = dag_run.conf or {}
    is_retry = dag_run_conf.get("retry_attempt", False)
    fallback_already_sent = dag_run_conf.get("fallback_email_already_sent", False)
    
    # For retry runs, use original_run_id for tracker key
    # For original failures, use current run_id
    if is_retry:
        original_run_id = dag_run_conf.get('original_run_id', run_id)
        original_dag_id = dag_run_conf.get('original_dag_id', dag_id)
        tracker_key = f"{original_dag_id}:{original_run_id}"
        
        logging.info("=" * 80)
        logging.info(f"🔄 RETRY RUN FAILED - Attempt failed for {tracker_key}")
        logging.info("=" * 80)
    else:
        # This is the ORIGINAL failure
        tracker_key = f"{dag_id}:{run_id}"
        original_run_id = run_id
        original_dag_id = dag_id
        
        logging.info("=" * 80)
        logging.info(f"❌ ORIGINAL FAILURE - Tracking {tracker_key}")
        logging.info("=" * 80)
    
    # ============================================================
    # STEP 2: Extract email_data (required for both original and retry)
    # ============================================================
    
    ti = context['ti']
    email_data = dag_run_conf.get("email_data")
    
    if not email_data:
        # Try XCom as fallback
        possible_keys = [
            "unread_emails",
            "no_action_emails",
            "llm_search_results",
            "reply_emails",
            "report_emails",
            "email_data"
        ]
        
        for key in possible_keys:
            data = ti.xcom_pull(key=key)
            if data:
                if isinstance(data, list) and len(data) > 0:
                    email_data = data[0]
                elif isinstance(data, dict):
                    email_data = data
                break
    
    if not email_data:
        logging.warning("Could not find email data in XCom for fallback - skipping email send")
        
        # 🔥 STILL TRACK THE FAILURE even without email_data
        retry_tracker = Variable.get("hubspot_retry_tracker", default_var={}, deserialize_json=True)
        
        retry_tracker[tracker_key] = {
            "status": "failed",
            "thread_id": "unknown",
            "failure_time": datetime.now(pytz.timezone("Asia/Kolkata")).isoformat(),
            "fallback_sent": False,
            "last_trigger_time": None,
            "email_data": None,
            "failed_task_id": task_id,
            "error": "No email_data available",
            "retry_count": 0,  # NEW: Track retry attempts
            "max_retries_reached": False  # NEW: Flag for max retries
        }
        
        Variable.set("hubspot_retry_tracker", json.dumps(retry_tracker))
        logging.warning(f"⚠️ Tracked failure WITHOUT email_data: {tracker_key}")
        return
    
    # Extract thread_id
    thread_id = email_data.get("threadId", email_data.get("thread_id"))
    
    if not thread_id:
        logging.warning("Could not determine thread_id for fallback")
        thread_id = "unknown"
    
    # ============================================================
    # STEP 3: Check retry count and max retries
    # ============================================================
    
    retry_tracker = Variable.get("hubspot_retry_tracker", default_var={}, deserialize_json=True)
    existing_entry = retry_tracker.get(tracker_key, {})
    current_retry_count = existing_entry.get("retry_count", 0)
    
    # NEW: Check if max retries already reached
    if current_retry_count >= 2:
        logging.warning(f"⚠️ MAX RETRIES REACHED ({current_retry_count}/2) for {tracker_key}")
        
        # Mark as permanently failed
        retry_tracker[tracker_key] = {
            "status": "max_retries_exceeded",
            "thread_id": thread_id,
            "failure_time": existing_entry.get("failure_time", datetime.now(pytz.timezone("Asia/Kolkata")).isoformat()),
            "fallback_sent": True,  # Ensure fallback is sent
            "last_trigger_time": existing_entry.get("last_trigger_time"),
            "email_data": email_data,
            "failed_task_id": task_id,
            "retry_count": current_retry_count,
            "max_retries_reached": True,
            "final_failure_time": datetime.now(pytz.timezone("Asia/Kolkata")).isoformat()
        }
        Variable.set("hubspot_retry_tracker", json.dumps(retry_tracker))
        
        # Send final fallback email if not already sent
        if not existing_entry.get("fallback_sent", False):
            send_email = True
        else:
            send_email = False
            logging.info(f"Final fallback already sent for {thread_id}")
    else:
        # Normal flow - check if fallback already sent for this thread
        send_email = False
        
        if is_retry and fallback_already_sent:
            logging.info(f"Retry run + fallback already sent → SKIP email for thread {thread_id}")
            send_email = False
        else:
            # Check fallback_sent_threads tracker
            try:
                fallback_tracker = ti.xcom_pull(
                    key="fallback_sent_threads",
                    dag_id=original_dag_id,
                    include_prior_dates=True
                ) or {}
                
                if thread_id in fallback_tracker:
                    logging.info(f"Fallback email already sent for thread {thread_id} - skipping duplicate")
                    send_email = False
                else:
                    send_email = True
            except Exception as e:
                logging.warning(f"Could not check fallback tracker: {e}")
                send_email = True  # Send by default if tracker check fails
    
    # ============================================================
    # STEP 4: Send fallback email if needed
    # ============================================================
    
    email_sent = False
    
    if send_email:
        headers = email_data.get("headers", {})
        sender = headers.get("From", "Unknown <unknown@email.com>")
        
        sender_name_match = re.search(r'^([^<]+)', sender)
        sender_name = sender_name_match.group(1).strip() if sender_name_match else "there"
        
        sender_email_match = re.search(r'<(.+?)>', sender)
        recipient_email = sender_email_match.group(1) if sender_email_match else None
        
        if not recipient_email:
            logging.error("Could not determine recipient email for fallback")
        else:
            service = authenticate_gmail()
            if service:
                # NEW: Customize message based on retry count
                if current_retry_count >= 2:
                    issue_message = "We've attempted to process your request multiple times, but continue to experience technical difficulties."
                    action_message = "Our team has been notified and will review your request manually. We'll reach out to you directly with the information you need."
                else:
                    issue_message = """We’re currently experiencing a temporary technical issue that may affect your experience with the Hubspot Email Companion. Our engineering team has already identified
the cause and is actively working on a resolution."""
                    action_message = """Our system will automatically retry your request within next 40 minutes, and we'll process it as soon as possible. In the meantime, your data and configurations remain
secure, and no action is required from your side. Thank you for your patience and understanding — we genuinely appreciate it."""
                
                fallback_html = f"""
<html>
<head>
    <style>
        body {{
            font-family: Arial, sans-serif;
            line-height: 1.6;
            color: #333;
            max-width: 600px;
            margin: 0 auto;
            padding: 20px;
        }}
        .greeting {{ margin-bottom: 15px; }}
        .message {{ margin: 15px 0; }}
        .closing {{ margin-top: 15px; }}
        .signature {{ margin-top: 15px; font-weight: bold; }}
        .company {{ color: #666; font-size: 0.9em; }}
    </style>
</head>
<body>
    <div class="greeting">
        <p>Hello {sender_name},</p>
    </div>
   
    <div class="message">
        <p>Thank you for your message to the HubSpot Assistant.</p>
       
        <p>{issue_message} Our engineering team has been notified and is actively working on a resolution.</p>
       
        <p>{action_message}</p>
    </div>
   
    <div class="closing">
        <p>If this is urgent, please feel free to reach out directly, and we'll assist you manually.</p>
    </div>
   
    <div class="signature">
        <p>Best regards,<br>
        The HubSpot Assistant Team<br>
        <a href="http://lowtouch.ai" class="company">lowtouch.ai</a></p>
    </div>
</body>
</html>
"""
                
                try:
                    original_message_id = headers.get("Message-ID", "")
                    references = headers.get("References", "")
                    if original_message_id and original_message_id not in references:
                        references = f"{references} {original_message_id}".strip()
                    
                    subject = headers.get("Subject", "No Subject")
                    if not subject.lower().startswith("re:"):
                        subject = f"Re: {subject}"

                    msg = MIMEMultipart()
                    msg["From"] = f"HubSpot via lowtouch.ai <{HUBSPOT_FROM_ADDRESS}>"
                    msg["To"] = recipient_email
                    msg["Subject"] = subject
                    if original_message_id:
                        msg["In-Reply-To"] = original_message_id
                        msg["References"] = references
                    
                    msg.attach(MIMEText(fallback_html, "html"))

                    raw_msg = base64.urlsafe_b64encode(msg.as_string().encode("utf-8")).decode("utf-8")
                    service.users().messages().send(
                        userId="me",
                        body={"raw": raw_msg}
                    ).execute()

                    email_sent = True
                    logging.info(f"✓ Fallback email sent successfully to {recipient_email} for thread {thread_id}")
                    
                    # Mark this thread as having received a fallback email
                    fallback_tracker[thread_id] = {
                        "timestamp": datetime.now(pytz.timezone("Asia/Kolkata")).isoformat(),
                        "recipient": recipient_email,
                        "dag_run_id": run_id,
                        "retry_count": current_retry_count
                    }
                    
                    ti.xcom_push(key="fallback_sent_threads", value=fallback_tracker)
                    
                except Exception as e:
                    logging.error(f"Failed to send fallback email to {recipient_email}: {e}", exc_info=True)
    
    # ============================================================
    # STEP 5: ALWAYS UPDATE RETRY TRACKER (critical fix)
    # ============================================================
    
    retry_tracker = Variable.get("hubspot_retry_tracker", default_var={}, deserialize_json=True)
    
    # Increment retry count for the next attempt
    new_retry_count = current_retry_count if is_retry else 0
    
    # Update or create tracker entry
    retry_tracker[tracker_key] = {
        "status": "max_retries_exceeded" if current_retry_count >= 3 else "failed",
        "thread_id": thread_id,
        "failure_time": existing_entry.get("failure_time", datetime.now(pytz.timezone("Asia/Kolkata")).isoformat()),
        "fallback_sent": email_sent or existing_entry.get("fallback_sent", False),
        "last_trigger_time": existing_entry.get("last_trigger_time"),
        "email_data": email_data,
        "failed_task_id": task_id,
        "is_retry_failure": is_retry,
        "retry_count": new_retry_count,  # NEW: Track attempts
        "max_retries_reached": current_retry_count >= 2,  # NEW: Flag
    }
    
    Variable.set("hubspot_retry_tracker", json.dumps(retry_tracker))
    
    if is_retry:
        logging.info(f"✓ Updated retry_tracker for retry failure: {tracker_key} (attempt {new_retry_count}/3)")
    else:
        logging.info(f"✓ Added original failure to retry_tracker: {tracker_key}")
    
    # Push failure metadata to XCom for tracking
    failure_metadata = {
        "dag_id": dag_id,
        "run_id": run_id,
        "task_id": task_id,
        "execution_date": str(execution_date),
        "failure_timestamp": datetime.now(pytz.timezone("Asia/Kolkata")).isoformat(),
        "callback_type": "fallback_email",
        "thread_id": thread_id,
        "email_sent": email_sent,
        "recipient_email": recipient_email if send_email else None,
        "fallback_already_sent": email_sent,
        "email_data": email_data,
        "is_retry_failure": is_retry,
        "tracker_key": tracker_key,
        "retry_count": new_retry_count,  # NEW
        "max_retries_reached": current_retry_count >= 2  # NEW
    }
    
    try:
        ti.xcom_push(key="failure_metadata", value=failure_metadata)
        logging.info(f"✓ Pushed failure metadata to XCom: {failure_metadata}")
    except Exception as e:
        logging.warning(f"Could not push failure metadata: {e}")

SLACK_WEBHOOK_URL = Variable.get("ltai.v3.hubspot.slack_webhook_url", default_var=None)
SERVER_NAME = Variable.get("ltai.v3.hubspot.server_name", default_var="UNKNOWN")

def send_hubspot_slack_alert(context):
    """
    Send Slack alert for HubSpot DAG failures.
    Tracks retry attempts and includes detailed error context.
    
    INTEGRATION POINTS:
    - Called on every task failure (via on_failure_callback)
    - Tracks retry attempt number (1/3, 2/3, 3/3)
    - Includes email thread ID and sender info
    - Shows which task failed and why
    """
    webhook_url = SLACK_WEBHOOK_URL
    if not webhook_url:
        logging.error("❌ Slack webhook URL not found in Airflow Variables")
        return

    dag_run = context.get("dag_run")
    task_instance = context.get("task_instance")
    
    dag_id = dag_run.dag_id
    run_id = dag_run.run_id
    task_id = task_instance.task_id
    execution_date = context.get("execution_date")
    
    # ============================================================
    # STEP 1: Determine if this is a retry or original failure
    # ============================================================
    
    dag_run_conf = dag_run.conf or {}
    is_retry = dag_run_conf.get("retry_attempt", False)
    
    if is_retry:
        original_run_id = dag_run_conf.get('original_run_id', run_id)
        original_dag_id = dag_run_conf.get('original_dag_id', dag_id)
        tracker_key = f"{original_dag_id}:{original_run_id}"
        retry_attempt_number = dag_run_conf.get('retry_attempt_number', 0)
    else:
        tracker_key = f"{dag_id}:{run_id}"
        original_run_id = run_id
        original_dag_id = dag_id
        retry_attempt_number = 1  # This is the first attempt
    
    # ============================================================
    # STEP 2: Get retry tracker info
    # ============================================================
    
    retry_tracker = Variable.get("hubspot_retry_tracker", default_var={}, deserialize_json=True)
    existing_entry = retry_tracker.get(tracker_key, {})
    current_retry_count = existing_entry.get("retry_count", 0)
    max_retries_reached = existing_entry.get("max_retries_reached", False)
    
    # ============================================================
    # STEP 3: Extract email context from conf or XCom
    # ============================================================
    
    ti = context['ti']
    email_data = dag_run_conf.get("email_data")
    
    if not email_data:
        # Try XCom as fallback
        possible_keys = [
            "unread_emails",
            "no_action_emails", 
            "llm_search_results",
            "reply_emails",
            "report_emails",
            "email_data"
        ]
        
        for key in possible_keys:
            data = ti.xcom_pull(key=key)
            if data:
                if isinstance(data, list) and len(data) > 0:
                    email_data = data[0]
                elif isinstance(data, dict):
                    email_data = data
                break
    
    # Extract thread and sender info
    thread_id = "unknown"
    sender_email = "unknown"
    subject = "No Subject"
    
    if email_data:
        thread_id = email_data.get("threadId", email_data.get("thread_id", "unknown"))
        headers = email_data.get("headers", {})
        sender_email = headers.get("From", "unknown")
        subject = headers.get("Subject", "No Subject")
    
    # ============================================================
    # STEP 4: Pull error message from XCom
    # ============================================================
    
    error_message = ti.xcom_pull(task_ids=task_id, key="error_message")
    if not error_message:
        # Try to get from exception
        try:
            exception_info = context.get('exception')
            if exception_info:
                error_message = str(exception_info)[:500]
        except:
            error_message = "No error details available"
    
    # ============================================================
    # STEP 5: Send final failure email to user if this is 3rd attempt
    # ============================================================
    
    if current_retry_count >= 2 and not existing_entry.get("final_fallback_sent", False):
        # This is the moment when 3rd retry just failed - send final email to user
        logging.info("🔴 Triggering final failure email to user...")
        try:
            send_final_failure_email_to_user(context)
        except Exception as email_error:
            logging.error(f"Failed to send final failure email: {email_error}")
    
    # ============================================================
    # STEP 6: Build Slack message with retry context
    # ============================================================
    
    # Determine alert severity and message style
    if max_retries_reached or current_retry_count >= 2:
        icon = "🔴"
        status = "FINAL FAILURE - MAX RETRIES EXCEEDED"
        color = "#ff0000"
        retry_text = f"❌ *Retry Attempt {current_retry_count}/2 FAILED*\n⚠️ *No more retry attempts will be made*\n📧 Manual intervention email sent to user"
    elif is_retry:
        icon = "🟡"
        status = f"RETRY FAILURE - Attempt {current_retry_count}/2"
        color = "#ffa500"
        retry_text = f"🔄 *Retry Attempt {current_retry_count}/2 Failed*\n⏰ Next retry in 20 minutes"
    else:
        icon = "🔴"
        status = "INITIAL FAILURE"
        color = "#ff4444"
        retry_text = "🔄 *Retry Attempt 1/2*\n⏰ Will retry in 20 minutes"
    
    # Build detailed message
    message_text = f"{icon} *HubSpot DAG Failure* - {status}"
    
    # Create rich Slack message with blocks
    slack_payload = {
    "text": f"HubSpot Task Failed - {SERVER_NAME}",  # fallback plain text
    "blocks": [
        {
            "type": "header",
            "text": {
                "type": "plain_text",
                "text": f"HubSpot Task Failed - {SERVER_NAME}",
                "emoji": False
            }
        },
        {
            "type": "section",
            "text": {
                "type": "mrkdwn",
                "text": f"*Status*\n{status}"
            }
        },
        {
            "type": "section",
            "text": {
                "type": "mrkdwn",
                "text": f"*DAG*\n{dag_id}"
            }
        },
        {
            "type": "section",
            "text": {
                "type": "mrkdwn",
                "text": f"*Task*\n{task_id}"
            }
        },
        {
            "type": "section",
            "text": {
                "type": "mrkdwn",
                "text": f"*Run ID*\n{run_id}..."
            }
        },
        {
            "type": "section",
            "text": {
                "type": "mrkdwn",
                "text": retry_text
            }
        },
        {
            "type": "divider"
        },
        {
            "type": "section",
            "text": {
                "type": "mrkdwn",
                "text": f"*Thread ID*\n{thread_id}..."
            }
        },
        {
            "type": "section",
            "text": {
                "type": "mrkdwn",
                "text": f"*Sender*\n{sender_email}"
            }
        },
        {
            "type": "section",
            "text": {
                "type": "mrkdwn",
                "text": f"*Subject*\n{subject}"
            }
        }
    ]
}

# Add error only if we have meaningful content
    if error_message and error_message.strip() and error_message != "No error details available":
        slack_payload["blocks"].append({
            "type": "section",
            "text": {
                "type": "mrkdwn",
                "text": f"*Error*\n{error_message.strip()[:800]}"
            }
        })

    # Optional: keep the color bar on the left for visual distinction
    slack_payload["attachments"] = [
        {
            "color": color,          # e.g. "#d0021b" for red/error, "#f5a623" for orange/warning
            "blocks": []             # empty → just shows the color stripe
        }
    ] 
    
    # ============================================================
    # STEP 7: Send to Slack
    # ============================================================
    
    try:
        response = requests.post(
            webhook_url,
            data=json.dumps(slack_payload),
            headers={"Content-Type": "application/json"},
            timeout=10,
        )
        
        if response.status_code == 200:
            logging.info(f"✓ Slack alert sent successfully for {tracker_key} (Attempt {current_retry_count}/2)")
        else:
            logging.error(f"❌ Slack alert failed: {response.status_code} - {response.text}")
            
    except Exception as e:
        logging.error(f"❌ Failed to send Slack alert: {e}")

def send_final_failure_email_to_user(context):
    """
    Send a final "manual intervention required" email to the user
    when all 3 retry attempts have been exhausted.
    
    This is ONLY called after the 3rd failed retry attempt.
    """
    dag_run = context.get('dag_run')
    task_instance = context.get('task_instance')
    
    dag_id = dag_run.dag_id
    run_id = dag_run.run_id
    task_id = task_instance.task_id
    
    # ============================================================
    # STEP 1: Check if this is the 3rd retry failure
    # ============================================================
    
    dag_run_conf = dag_run.conf or {}
    is_retry = dag_run_conf.get("retry_attempt", False)
    
    if is_retry:
        original_run_id = dag_run_conf.get('original_run_id', run_id)
        original_dag_id = dag_run_conf.get('original_dag_id', dag_id)
        tracker_key = f"{original_dag_id}:{original_run_id}"
    else:
        tracker_key = f"{dag_id}:{run_id}"
    
    # Get retry tracker
    retry_tracker = Variable.get("hubspot_retry_tracker", default_var={}, deserialize_json=True)
    existing_entry = retry_tracker.get(tracker_key, {})
    current_retry_count = existing_entry.get("retry_count", 0)
    
    # ============================================================
    # STEP 2: Only proceed if this is the final (3rd) failure
    # ============================================================
    
    if current_retry_count < 2:
        logging.info(f"Not final failure yet (attempt {current_retry_count}/2) - skipping final email")
        return
    
    logging.info(f"🔴 FINAL FAILURE DETECTED - Sending manual intervention email")
    
    # ============================================================
    # STEP 3: Extract email data
    # ============================================================
    
    ti = context['ti']
    email_data = dag_run_conf.get("email_data")
    
    if not email_data:
        # Try XCom as fallback
        possible_keys = [
            "unread_emails",
            "no_action_emails",
            "llm_search_results",
            "reply_emails",
            "report_emails",
            "email_data"
        ]
        
        for key in possible_keys:
            data = ti.xcom_pull(key=key)
            if data:
                if isinstance(data, list) and len(data) > 0:
                    email_data = data[0]
                elif isinstance(data, dict):
                    email_data = data
                break
    
    if not email_data:
        logging.warning("Could not find email data for final failure email")
        return
    
    # Extract sender info
    headers = email_data.get("headers", {})
    sender = headers.get("From", "Unknown <unknown@email.com>")
    
    sender_name_match = re.search(r'^([^<]+)', sender)
    sender_name = sender_name_match.group(1).strip() if sender_name_match else "there"
    
    sender_email_match = re.search(r'<(.+?)>', sender)
    recipient_email = sender_email_match.group(1) if sender_email_match else None
    
    if not recipient_email:
        logging.error("Could not determine recipient email for final failure email")
        return
    
    # ============================================================
    # STEP 4: Build the "manual intervention" email
    # ============================================================
    
    service = authenticate_gmail()
    if not service:
        logging.error("Gmail authentication failed - cannot send final failure email")
        return
    
    final_failure_html = f"""
<html>
<head>
    <style>
        body {{
            font-family: Arial, sans-serif;
            line-height: 1.6;
            color: #333;
            max-width: 600px;
            margin: 0 auto;
            padding: 20px;
        }}
        .greeting {{ margin-bottom: 15px; }}
        .message {{ margin: 15px 0; }}
        .closing {{ margin-top: 15px; }}
        .signature {{ margin-top: 15px; font-weight: bold; }}
        .company {{ color: #666; font-size: 0.9em; }}
    </style>
</head>
<body>
    <div class="greeting">
        <p>Hello {sender_name},</p>
    </div>
   
    <div class="message">
        <p>We’re still experiencing a technical issue, and the maximum retry attempts for your request have been exceeded.</p>
        
        <p>To avoid further delays, please proceed with the required action manually in your HubSpot account.</p>
    </div>
   
    <div class="closing">
        <p>If you need assistance, please reply to this email or contact your HubSpot administrator.</p>
    </div>
   
    <div class="signature">
        <p>Best regards,<br>
        The HubSpot Assistant Team<br>
        <a href="http://lowtouch.ai" class="company">lowtouch.ai</a></p>
    </div>
</body>
</html>
"""

    
    # Build the message
    msg = MIMEMultipart()
    msg["From"] = f"HubSpot via lowtouch.ai <{HUBSPOT_FROM_ADDRESS}>"
    msg["To"] = recipient_email
    
    # Set subject as a reply
    subject = headers.get("Subject", "Manual Action Required")
    if not subject.lower().startswith("re:"):
        subject = f"Re: {subject}"
    msg["Subject"] = subject
    
    # NEW: Add threading headers to keep it in the same thread
    original_message_id = headers.get("Message-ID", "")
    references = headers.get("References", "")
    if original_message_id:
        msg["In-Reply-To"] = original_message_id
        if original_message_id not in references:
            references = f"{references} {original_message_id}".strip()
        msg["References"] = references
    
    msg.attach(MIMEText(final_failure_html, "html"))
    
    # Send the email
    raw_msg = base64.urlsafe_b64encode(msg.as_string().encode("utf-8")).decode("utf-8")
    try:
        service.users().messages().send(
            userId="me",
            body={"raw": raw_msg}
        ).execute()
        logging.info(f"✓ Final failure email sent successfully to {recipient_email} in thread")
        
        # Update tracker to mark final fallback as sent
        existing_entry["final_fallback_sent"] = True
        retry_tracker[tracker_key] = existing_entry
        Variable.set("hubspot_retry_tracker", json.dumps(retry_tracker))
    except Exception as e:
        logging.error(f"Failed to send final failure email: {e}", exc_info=True)

def clear_retry_tracker_on_success(context):
    dag_run = context['dag_run']
    original_run_id = dag_run.conf.get('original_run_id')
    original_dag_id = dag_run.conf.get('original_dag_id', dag_run.dag_id)
    
    if not original_run_id or not dag_run.conf.get('retry_attempt'):
        return  # Not a retry run
    
    tracker_key = f"{original_dag_id}:{original_run_id}"
    
    retry_tracker = Variable.get("hubspot_retry_tracker", default_var={}, deserialize_json=True)
    
    if tracker_key in retry_tracker:
        del retry_tracker[tracker_key]
        Variable.set("hubspot_retry_tracker", json.dumps(retry_tracker))
        logging.info(f"Retry succeeded - cleared tracker for {tracker_key}")

def update_retry_tracker_on_failure(context):
    dag_run = context['dag_run']
    original_run_id = dag_run.conf.get('original_run_id')
    original_dag_id = dag_run.conf.get('original_dag_id', dag_run.dag_id)
    
    if not original_run_id or not dag_run.conf.get('retry_attempt'):
        return  # Not a retry run
    
    tracker_key = f"{original_dag_id}:{original_run_id}"
    
    retry_tracker = Variable.get("hubspot_retry_tracker", default_var={}, deserialize_json=True)
    
    if tracker_key in retry_tracker:
        retry_tracker[tracker_key]["status"] = "failed"
        Variable.set("hubspot_retry_tracker", json.dumps(retry_tracker))
        logging.info(f"Retry failed - updated status for {tracker_key}")

default_args = {
    "owner": "lowtouch.ai_developers",
    "depends_on_past": False,
    "start_date": datetime(2025, 8, 22),
    "retry_delay": timedelta(seconds=15),
    'on_failure_callback': [
        send_fallback_email_on_failure,  # Sends email to user
        send_hubspot_slack_alert          # Sends Slack alert to team
    ]
}

@task(task_id="handle_retry_if_present")
def handle_retry_or_normal(**context):
    conf = context["dag_run"].conf
    
    if conf.get("retry_attempt", False):
        email_data = conf.get("email_data")
        thread_id = conf.get("thread_id")
        
        # Very important: reset fallback flag for this thread!
        ti = context["ti"]
        tracker = ti.xcom_pull(key="fallback_sent_threads", include_prior_dates=True) or {}
        if thread_id in tracker:
            del tracker[thread_id]  # ← remove so fallback can be sent again if needed
            ti.xcom_push(key="fallback_sent_threads", value=tracker)
        
        # You can also skip fetch_unread_emails completely
        return {
            "unread_emails": [email_data],  # ← inject directly!
            "is_retry": True
        }
    
    return {"is_retry": False}

HUBSPOT_FROM_ADDRESS = Variable.get("ltai.v3.hubspot.from.address")
GMAIL_CREDENTIALS = Variable.get("ltai.v3.hubspot.gmail.credentials")
LAST_PROCESSED_EMAIL_FILE = "/appz/cache/hubspot_last_processed_email.json"
OLLAMA_HOST = Variable.get("ltai.v3.hubspot.ollama.host","http://agentomatic:8000")
HUBSPOT_API_KEY = Variable.get("ltai.v3.husbpot.api.key")
HUBSPOT_BASE_URL = Variable.get("ltai.v3.hubspot.url")
HUBSPOT_UI_URL = Variable.get("ltai.v3.hubspot.ui.url")

def authenticate_gmail():
    try:
        creds = Credentials.from_authorized_user_info(json.loads(GMAIL_CREDENTIALS))
        service = build("gmail", "v1", credentials=creds)
        profile = service.users().getProfile(userId="me").execute()
        logged_in_email = profile.get("emailAddress", "")
        if logged_in_email.lower() != HUBSPOT_FROM_ADDRESS.lower():
            raise ValueError(f"Wrong Gmail account! Expected {HUBSPOT_FROM_ADDRESS}, but got {logged_in_email}")
        logging.info(f"Authenticated Gmail account: {logged_in_email}")
        return service
    except Exception as e:
        logging.error(f"Failed to authenticate Gmail: {str(e)}")
        return None

def get_last_checked_timestamp():
    if os.path.exists(LAST_PROCESSED_EMAIL_FILE):
        with open(LAST_PROCESSED_EMAIL_FILE, "r") as f:
            last_checked = json.load(f).get("last_processed", None)
            if last_checked:
                return last_checked
    current_timestamp_ms = int(time.time() * 1000)
    update_last_checked_timestamp(current_timestamp_ms)
    return current_timestamp_ms

def update_last_checked_timestamp(timestamp):
    os.makedirs(os.path.dirname(LAST_PROCESSED_EMAIL_FILE), exist_ok=True)
    with open(LAST_PROCESSED_EMAIL_FILE, "w") as f:
        json.dump({"last_processed": timestamp}, f)

def decode_email_payload(msg):
    """Decode email payload to extract text content."""
    try:
        if msg.is_multipart():
            for part in msg.walk():
                content_type = part.get_content_type()
                if content_type in ["text/plain", "text/html"]:
                    try:
                        return part.get_payload(decode=True).decode()
                    except UnicodeDecodeError:
                        return part.get_payload(decode=True).decode('latin-1')
        else:
            try:
                return msg.get_payload(decode=True).decode()
            except UnicodeDecodeError:
                return msg.get_payload(decode=True).decode('latin-1')
        return ""
    except Exception as e:
        logging.error(f"Error decoding email payload: {e}")
        return ""

def is_email_authorized(raw_email: str) -> bool:
    """
    Returns True if the full email address is authorized.
    Authorization = present in EITHER whitelist (full address match only).
    """
    try:
        # Load both whitelists once
        company_raw = Variable.get("ltai.v3.hubspot.email.whitelist", default_var="[]")
        external_raw = Variable.get("hubspot.email.whitelist.external", default_var="[]")
        
        company_list = json.loads(company_raw)
        external_list = json.loads(external_raw)
        
        # Combine into a single set of lowercase authorized emails
        authorized_emails = {e.lower() for e in company_list + external_list}
        
        # Extract clean email from "Name <email@domain.com>" or plain email
        match = re.search(r'<([^>]+)>', raw_email)
        clean_email = (match.group(1) if match else raw_email).strip().lower()
        
        # Basic validation (optional but recommended)
        if '@' not in clean_email or len(clean_email) < 5:
            logging.warning(f"Invalid email format treated as unauthorized: {raw_email}")
            return False
        
        is_auth = clean_email in authorized_emails
        
        logging.info(
            f"Authorization check — "
            f"Email: {clean_email} | "
            f"In company list: {clean_email in {e.lower() for e in company_list}} | "
            f"In external list: {clean_email in {e.lower() for e in external_list}} | "
            f"Final: {'AUTHORIZED' if is_auth else 'NOT AUTHORIZED'}"
        )
        
        return is_auth
        
    except json.JSONDecodeError as e:
        logging.error(f"Whitelist JSON parse error: {e} → treating as unauthorized")
        return False
    except Exception as e:
        logging.error(f"Unexpected error in authorization check: {e}")
        return False

def get_email_thread(service, email_data):
    """
    Retrieve full email thread by reconstructing from References header and searching for each message.
    """
    try:
        if not email_data or "headers" not in email_data:
            logging.error("Invalid email_data: 'headers' key missing")
            return []

        thread_id = email_data.get("threadId")
        headers = email_data.get("headers", {})
        current_message_id = headers.get("Message-ID", "")
        email_id = email_data.get("id", "")
        
        logging.info(f"Processing email: email_id={email_id}, thread_id={thread_id}")

        # ✅ Extract all message IDs from References header
        references = headers.get("References", "")
        in_reply_to = headers.get("In-Reply-To", "")
        
        # Parse message IDs from References
        message_ids = []
        if references:
            # References contains space-separated Message-IDs
            message_ids = re.findall(r'<([^>]+)>', references)
        if in_reply_to:
            reply_to_id = re.search(r'<([^>]+)>', in_reply_to)
            if reply_to_id and reply_to_id.group(1) not in message_ids:
                message_ids.append(reply_to_id.group(1))
        
        # Add current message ID
        current_id = re.search(r'<([^>]+)>', current_message_id)
        if current_id:
            message_ids.append(current_id.group(1))
        
        logging.info(f"Found {len(message_ids)} message IDs in thread from References header")
        
        # Now fetch each message by searching for its Message-ID
        processed_thread = []
        for msg_id in message_ids:
            try:
                # Search for message by RFC822 Message-ID
                search_query = f'rfc822msgid:{msg_id}'
                search_result = service.users().messages().list(
                    userId="me",
                    q=search_query,
                    maxResults=1
                ).execute()
                
                messages = search_result.get("messages", [])
                if not messages:
                    logging.warning(f"Could not find message with ID: {msg_id}")
                    continue
                
                # Fetch the full message
                gmail_msg_id = messages[0]["id"]
                raw_message = service.users().messages().get(
                    userId="me",
                    id=gmail_msg_id,
                    format="raw"
                ).execute()
                
                raw_msg = base64.urlsafe_b64decode(raw_message["raw"])
                email_msg = message_from_bytes(raw_msg)
                
                # Get metadata for headers
                metadata = service.users().messages().get(
                    userId="me",
                    id=gmail_msg_id,
                    format="metadata",
                    metadataHeaders=["From", "Subject", "Date", "Message-ID", "In-Reply-To", "References", "X-Task-ID", "X-Task-Type"]
                ).execute()
                
                msg_headers = {}
                for h in metadata.get("payload", {}).get("headers", []):
                    msg_headers[h["name"]] = h["value"]
                
                content = decode_email_payload(email_msg)
                from_address = msg_headers.get("From", "").lower()
                is_from_bot = HUBSPOT_FROM_ADDRESS.lower() in from_address
                timestamp = int(metadata.get("internalDate", 0))

                processed_thread.append({
                    "headers": msg_headers,
                    "content": content.strip(),
                    "timestamp": timestamp,
                    "from_bot": is_from_bot,
                    "message_id": gmail_msg_id,
                    "role": "assistant" if is_from_bot else "user"
                })
                
                logging.info(f"✓ Retrieved message {len(processed_thread)}/{len(message_ids)}: {msg_id[:30]}...")
                
            except Exception as e:
                logging.error(f"Error fetching message {msg_id}: {e}")
                continue

        # Sort by timestamp
        processed_thread.sort(key=lambda x: x.get("timestamp", 0))
        
        logging.info(f"✓ Processed thread {thread_id} with {len(processed_thread)} messages")
        for idx, email in enumerate(processed_thread, 1):
            role = email['role']
            from_email = email['headers'].get('From', 'Unknown')[:40]
            preview = email['content'][:60].replace('\n', ' ')
            timestamp = email['timestamp']
            logging.info(f"  [{idx}] {role} @ {timestamp}: from={from_email}, preview={preview}...")

        return processed_thread

    except Exception as e:
        logging.error(f"Error retrieving thread: {e}", exc_info=True)
        return []

def format_chat_history(thread_history):
    """
    Convert thread history to chat history format compatible with the agent.
    Format matches agent.py structure with HumanMessage and AIMessage.
    
    Args:
        thread_history: List of email messages with role, content, etc.
    
    Returns:
        List of formatted messages for agent consumption
    """
    chat_history = []
    
    for msg in thread_history[:-1]:  # Exclude the latest message (will be sent as current prompt)
        # Create message in format similar to agent.py
        message = {
            "role": msg["role"],  # "user" or "assistant"
            "content": msg["content"]
        }
        chat_history.append(message)
    
    logging.info(f"Formatted chat history with {len(chat_history)} messages")
    return chat_history


def extract_all_recipients(email_data):
    """Extract all recipients (To, Cc, Bcc) from email headers."""
    headers = email_data.get("headers", {})
    
    def parse_addresses(header_value):
        """Parse email addresses from header string"""
        if not header_value:
            return []
        
        addresses = []
        for addr in header_value.split(','):
            addr = addr.strip()
            email_match = re.search(r'<([^>]+)>', addr)
            if email_match:
                addresses.append(email_match.group(1).strip())
            elif '@' in addr:
                addresses.append(addr.strip())
        
        return addresses
    
    to_recipients = parse_addresses(headers.get("To", ""))
    cc_recipients = parse_addresses(headers.get("Cc", ""))
    bcc_recipients = parse_addresses(headers.get("Bcc", ""))
    
    return {
        "to": to_recipients,
        "cc": cc_recipients,
        "bcc": bcc_recipients
    }

def send_email_reply(service, email, html_content):
    """Exact same sending logic as your original code."""
    headers = email.get("headers", {})
    email_id = email.get("id", "unknown")

    try:
        # Threading
        original_message_id = headers.get("Message-ID", "")
        references = headers.get("References", "")
        if original_message_id:
            references = f"{references} {original_message_id}".strip() if references else original_message_id

        # Subject
        subject = headers.get("Subject", "No Subject")
        if not subject.lower().startswith("re:"):
            subject = f"Re: {subject}"

        # Recipients
        all_recipients = extract_all_recipients(email)
        sender_email = headers.get("From", "")

        primary_recipient = sender_email

        cc_recipients = [
            addr for addr in all_recipients["to"] + all_recipients["cc"]
            if addr.lower() != sender_email.lower()
            and HUBSPOT_FROM_ADDRESS.lower() not in addr.lower()
        ]
        bcc_recipients = [
            addr for addr in all_recipients["bcc"]
            if HUBSPOT_FROM_ADDRESS.lower() not in addr.lower()
        ]

        cc_string = ', '.join(cc_recipients) if cc_recipients else None
        bcc_string = ', '.join(bcc_recipients) if bcc_recipients else None

        # Compose
        msg = MIMEMultipart()
        msg["From"] = f"HubSpot via lowtouch.ai <{HUBSPOT_FROM_ADDRESS}>"
        msg["To"] = primary_recipient
        if cc_string: msg["Cc"] = cc_string
        if bcc_string: msg["Bcc"] = bcc_string
        msg["Subject"] = subject
        if original_message_id: msg["In-Reply-To"] = original_message_id
        if references: msg["References"] = references
        msg.attach(MIMEText(html_content, "html"))

        # Send
        raw_msg = base64.urlsafe_b64encode(msg.as_string().encode("utf-8")).decode("utf-8")
        service.users().messages().send(
            userId="me",
            body={"raw": raw_msg}
        ).execute()

        logging.info(f"Sent reply to {sender_email} for email {email_id}")
        return True

    except Exception as e:
        logging.error(f"Failed to send email {email_id}: {e}", exc_info=True)
        return False

def check_if_task_completion_reply(email_data):
    # Check direct headers (for non-replies)
    headers = email_data.get("headers", {})
    task_id = headers.get("X-Task-ID")
    task_type = headers.get("X-Task-Type")
    if task_id and task_type == "daily-reminder":
        logging.info(f"✓ Detected task completion reply (direct headers) for task {task_id}")
        return task_id

    # If reply, check thread for bot's message
    if email_data.get("is_reply", False):
        thread_history = email_data.get("thread_history", [])
        for msg in reversed(thread_history):  # Newest to oldest
            if msg.get("from_bot", False):
                msg_headers = msg.get("headers", {})
                task_id = msg_headers.get("X-Task-ID")
                task_type = msg_headers.get("X-Task-Type")
                if task_id and task_type == "daily-reminder":
                    logging.info(f"✓ Detected task completion reply via thread for task {task_id}")
                    return task_id
                break  # Assume latest bot message is the relevant one

    return None

def trigger_task_completion_dag(**kwargs):
    """Trigger the task completion handler DAG"""
    ti = kwargs['ti']
    task_completion_emails = ti.xcom_pull(task_ids="branch_task", key="task_completion_emails") or []
    
    if not task_completion_emails:
        logging.info("No task completion emails to process")
        return
    
    for email in task_completion_emails:
        task_id = email.get("task_id")
        
        trigger_conf = {
            "email_data": email,
            "task_id": task_id,
            "chat_history": email.get("chat_history", []),
            "thread_history": email.get("thread_history", []),
            "thread_id": email.get("threadId", ""),
            "message_id": email.get("id", "")
        }
        
        task_id_safe = task_id.replace('-', '_')
        trigger_task = TriggerDagRunOperator(
            task_id=f"trigger_task_completion_{task_id_safe}",
            trigger_dag_id="hubspot_task_completion_handler",
            conf=trigger_conf,
        )
        trigger_task.execute(context=kwargs)
        
        logging.info(f"✓ Triggered task completion DAG for task {task_id}")
    
    logging.info(f"Triggered task completion handler for {len(task_completion_emails)} emails")

def fetch_unread_emails(**kwargs):
    """
    Fetch unread emails from Gmail OR use injected data in retry mode.
    Must run AFTER the 'handle_retry_or_normal' task.
    """
    ti = kwargs['ti']
    dag_run = kwargs['dag_run']
    
    # ──────────────────────────────────────────────────────────────
    # Step 1: Check if this is a retry run
    # ──────────────────────────────────────────────────────────────
    if dag_run.conf.get("retry_attempt", False):
        logging.info("=" * 80)
        logging.info("🔄 RETRY MODE - Using Stored Email Data from Failure")
        logging.info("=" * 80)
        
        # Get email data from the retry conf (stored during failure)
        email_data = dag_run.conf.get("email_data")
        thread_id = dag_run.conf.get("thread_id")
        
        if not email_data:
            logging.error("❌ RETRY FAILED: No email_data in retry conf!")
            logging.error(f"Available conf keys: {list(dag_run.conf.keys())}")
            ti.xcom_push(key="unread_emails", value=[])
            return []
        
        logging.info(f"✓ Retrieved stored email data for retry")
        logging.info(f"  Email ID: {email_data.get('id', 'unknown')}")
        logging.info(f"  Thread ID: {thread_id or email_data.get('threadId', 'unknown')}")
        logging.info(f"  Subject: {email_data.get('headers', {}).get('Subject', 'No Subject')}")
        logging.info(f"  From: {email_data.get('headers', {}).get('From', 'Unknown')}")
        
        # Reset fallback flag for this thread
        tracker = ti.xcom_pull(key="fallback_sent_threads", include_prior_dates=True) or {}
        if thread_id and thread_id in tracker:
            logging.info(f"✓ Resetting fallback flag for thread {thread_id}")
            del tracker[thread_id]
            ti.xcom_push(key="fallback_sent_threads", value=tracker)
        
        # Add retry metadata
        email_data["is_retry"] = True
        email_data["retry_count"] = email_data.get("retry_count", 0) + 1
        email_data["original_failure_task"] = dag_run.conf.get("original_task_id", "unknown")
        email_data["retry_timestamp"] = dag_run.conf.get("retry_timestamp")
        
        injected_emails = [email_data]
        ti.xcom_push(key="unread_emails", value=injected_emails)
        
        logging.info("=" * 80)
        return injected_emails
    
    # ──────────────────────────────────────────────────────────────
    # Step 2: Normal mode - fetch fresh unread emails from Gmail
    # ──────────────────────────────────────────────────────────────
    logging.info("NORMAL MODE → fetching unread emails from Gmail")
    
    service = authenticate_gmail()
    if not service:
        logging.error("Gmail authentication failed, skipping email fetch.")
        ti.xcom_push(key="unread_emails", value=[])
        return []
    
    last_checked_timestamp = get_last_checked_timestamp()
    last_checked_seconds = (
        last_checked_timestamp // 1000
        if last_checked_timestamp > 1000000000000
        else last_checked_timestamp
    )
    
    query = f"is:unread after:{last_checked_seconds}"
    logging.info(f"Fetching emails with query: {query}")
    
    try:
        results = service.users().messages().list(
            userId="me",
            labelIds=["INBOX"],
            q=query
        ).execute()
        messages = results.get("messages", [])
    except Exception as e:
        logging.error(f"Error fetching messages: {e}")
        ti.xcom_push(key="unread_emails", value=[])
        return []
    
    unread_emails = []
    max_timestamp = last_checked_timestamp
    processed_message_ids = set()
    
    logging.info(f"Found {len(messages)} unread messages to process")
    
    for msg in messages:
        msg_id = msg["id"]
        
        if msg_id in processed_message_ids:
            logging.info(f"Skipping already processed message: {msg_id}")
            continue
            
        try:
            msg_data = service.users().messages().get(
                userId="me",
                id=msg_id,
                format="metadata",
                metadataHeaders=[
                    "From", "To", "Cc", "Bcc", "Subject", "Date",
                    "Message-ID", "References", "In-Reply-To",
                    "X-Task-ID", "X-Task-Type"
                ]
            ).execute()
        except Exception as e:
            logging.error(f"Error fetching message {msg_id}: {e}")
            continue
            
        headers = {h["name"]: h["value"] for h in msg_data["payload"]["headers"]}
        sender = headers.get("From", "")
        
        # === FINAL: AUTHORIZATION CHECK (From + To + CC) ===
        # EXCLUDE BOT'S OWN ADDRESS ENTIRELY FROM CHECKS
        recipients = extract_all_recipients({"headers": headers})
        addresses_to_check = [sender] + recipients["to"] + recipients["cc"]
        
        # Clean and deduplicate raw addresses
        unique_raw_addresses = list(set(a.strip() for a in addresses_to_check if a.strip()))
        
        evaluated_clean_emails = []
        authorized = True
        
        for raw_addr in unique_raw_addresses:
            try:
                # Extract clean email
                match = re.search(r'<([^>]+)>', raw_addr)
                clean_email = (match.group(1).strip() if match else raw_addr.strip()).lower()
                
                # Basic malformed check
                if '@' not in clean_email or '.' not in clean_email.split('@')[-1]:
                    logging.warning(f"Malformed email treated as unauthorized: {raw_addr}")
                    authorized = False
                    evaluated_clean_emails.append(clean_email)
                    continue
                
                # === EXCLUDE BOT'S OWN ADDRESS FROM AUTH CHECK ===
                if clean_email == HUBSPOT_FROM_ADDRESS.lower():
                    logging.info(f"Skipping authorization check — this is the bot's own address: {clean_email}")
                    evaluated_clean_emails.append(clean_email)
                    continue  # Do not fail authorization because of bot address
                
                # === CHECK USER ADDRESSES ONLY ===
                if not is_email_authorized(raw_addr):
                    authorized = False
                
                evaluated_clean_emails.append(clean_email)
                
            except Exception as e:
                logging.warning(f"Error parsing address {raw_addr}: {e} → treated as unauthorized")
                authorized = False
        
        # === REQUIRED LOGGING (Acceptance Criteria) ===
        logging.info(f"Evaluated raw addresses: {', '.join(unique_raw_addresses)}")
        logging.info(f"Cleaned evaluated emails: {', '.join(evaluated_clean_emails)}")
        logging.info(f"Final authorization decision: {'authorized' if authorized else 'not authorized'}")
        
        if not authorized:
            logging.info(f"Unauthorized email (msg_id={msg_id}) → silently skipping all processing")
            mark_message_as_read(service, msg_id)
            continue  # Stops thread fetch, AI, DAG triggers, etc.
        # === END AUTHORIZATION CHECK ===
        
        sender = sender.lower()
        timestamp = int(msg_data["internalDate"])
        thread_id = msg_data.get("threadId", "")
        
        logging.info(f"Processing message ID: {msg_id}, From: {sender}, Timestamp: {timestamp}, Thread ID: {thread_id}")
        
        # Skip old messages
        if timestamp <= last_checked_timestamp:
            logging.info(f"Skipping old message {msg_id} - timestamp {timestamp} <= {last_checked_timestamp}")
            continue
        
        # Skip no-reply emails
        if "no-reply" in sender or "noreply" in sender:
            logging.info(f"Skipping no-reply email from: {sender}")
            continue
        
        # Skip emails from bot itself
        if sender == HUBSPOT_FROM_ADDRESS.lower():
            logging.info(f"Skipping email from bot: {sender}")
            continue
        logging.info(f"Processing message ID: {msg_id}, From: {sender}, Timestamp: {timestamp}, Thread ID: {thread_id}")
        
        
        # ✅ Construct email_object with all headers INCLUDING recipient headers
        email_object = {
            "id": msg_id,
            "threadId": thread_id,
            "headers": headers,
            "content": "",
            "timestamp": timestamp,
            "internalDate": timestamp
        }

        # ⭐ Extract all recipients from this email
        all_recipients = extract_all_recipients(email_object)
        
        logging.info(f"Extracted recipients - To: {all_recipients.get('to', [])}, "
                    f"Cc: {all_recipients.get('cc', [])}, Bcc: {all_recipients.get('bcc', [])}")
        
        # Get thread history
        thread_history = get_email_thread(service, email_object)
        
        if not thread_history:
            logging.error(f"Failed to retrieve thread history for message {msg_id}")
            continue
        
        # Format chat history (all messages except the latest)
        chat_history = format_chat_history(thread_history)
        
        # Extract latest message content (current user prompt)
        latest_message = thread_history[-1]
        latest_message_content = extract_latest_reply(latest_message["content"])
        
        # Update email_object with the latest message content
        email_object["content"] = latest_message_content
        
        # Determine if this is a reply based on thread length and email headers
        is_reply = len(thread_history) > 1
        subject = headers.get("Subject", "")
        is_reply = is_reply or subject.lower().startswith("re:") or bool(headers.get("In-Reply-To")) or bool(headers.get("References"))
        
        # ✅ Add all required fields to email_object including recipients
        email_object.update({
            "is_reply": is_reply,
            "thread_history": thread_history,
            "chat_history": chat_history,
            "thread_length": len(thread_history),
            "all_recipients": all_recipients
        })
        
        logging.info(f"Email {msg_id}: is_reply={is_reply}, thread_length={len(thread_history)}, "
                    f"chat_history_length={len(chat_history)}, recipients={len(all_recipients.get('to', [])) + len(all_recipients.get('cc', [])) + len(all_recipients.get('bcc', []))}")
        
        unread_emails.append(email_object)
        processed_message_ids.add(msg_id)
        mark_message_as_read(service, msg_id)
        if timestamp > max_timestamp:
            max_timestamp = timestamp
            
    # Update timestamp
    if messages:
        update_timestamp = max_timestamp + 1
        update_last_checked_timestamp(update_timestamp)
        logging.info(f"Updated last checked timestamp to: {update_timestamp}")
    else:
        logging.info("No messages found, timestamp unchanged")
        
    kwargs['ti'].xcom_push(key="unread_emails", value=unread_emails)
    logging.info(f"Processed {len(unread_emails)} emails ({sum(1 for e in unread_emails if e['is_reply'])} replies)")
    
    return unread_emails

def mark_message_as_read(service, message_id):
    try:
        service.users().messages().modify(
            userId='me',
            id=message_id,
            body={'removeLabelIds': ['UNREAD']}
        ).execute()
        logging.info(f"Marked message {message_id} as read")
        return True
    except Exception as e:
        logging.error(f"Error marking message {message_id} as read: {e}")
        return False
    
    try:
        # Reuse your existing send function
        send_email_reply(
            service=service,
            email={"headers": headers},  # minimal structure needed
            html_content=fallback_html
        )
        logging.info(f"✓ Fallback email sent to {recipient_email} after task failure")
        
    except Exception as e:
        logging.error(f"Failed to send fallback email to {recipient_email}: {e}")

def get_ai_response(prompt, conversation_history=None, expect_json=False, stream=True, max_retries=3):
    """
    Get AI response with robust error handling and retry logic.
    
    Args:
        prompt: The user prompt
        conversation_history: Previous conversation context
        expect_json: Whether to expect JSON response
        stream: Whether to use streaming mode
        max_retries: Maximum number of retry attempts
    """
    
    for attempt in range(max_retries):
        try:
            client = Client(
                host=OLLAMA_HOST, 
                headers={'x-ltai-client': 'hubspot-v6af'},
                timeout=60.0  # Add explicit timeout
            )
            messages = []

            if expect_json:
                messages.append({
                    "role": "system",
                    "content": "You are a JSON-only API. Always respond with valid JSON objects. Never include explanatory text, HTML, or markdown formatting. Only return the requested JSON structure."
                })

            if conversation_history:
                for item in conversation_history:
                    messages.append({"role": "user", "content": item["prompt"]})
                    messages.append({"role": "assistant", "content": item["response"]})
            
            messages.append({"role": "user", "content": prompt})
            
            # Add retry-specific logging
            if attempt > 0:
                logging.info(f"Retry attempt {attempt + 1}/{max_retries}")
            
            response = client.chat(
                model='hubspot:v6af', 
                messages=messages, 
                stream=stream,
                options={
                    'num_predict': 2048,  # Limit response length to reduce timeout risk
                }
            )
            
            # Handle response based on streaming mode
            if stream:
                ai_content = ""
                chunk_count = 0
                last_chunk_time = time.time()
                
                try:
                    for chunk in response:
                        chunk_count += 1
                        current_time = time.time()
                        
                        # Check for stalled stream (no chunks for 30 seconds)
                        if current_time - last_chunk_time > 120:
                            logging.warning(f"Stream stalled after {chunk_count} chunks")
                            raise TimeoutError("Stream stalled - no chunks received for 120 seconds")
                        
                        last_chunk_time = current_time
                        
                        if hasattr(chunk, 'message') and hasattr(chunk.message, 'content'):
                            ai_content += chunk.message.content
                        else:
                            logging.warning(f"Chunk {chunk_count} lacks expected structure")
                    
                    logging.info(f"Successfully received {chunk_count} chunks")
                    
                except (ConnectionError, TimeoutError, Exception) as stream_error:
                    logging.error(f"Streaming error on attempt {attempt + 1}: {stream_error}")
                    
                    # If we got partial content and it's the last attempt, use what we have
                    if ai_content.strip() and attempt == max_retries - 1:
                        logging.warning("Using partial response from failed stream")
                    else:
                        # Retry with non-streaming mode on next attempt
                        if attempt < max_retries - 1:
                            logging.info("Retrying with non-streaming mode")
                            time.sleep(2 ** attempt)  # Exponential backoff
                            stream = False  # Switch to non-streaming for retry
                            continue
                        raise
                        
            else:
                # Non-streaming mode
                if not (hasattr(response, 'message') and hasattr(response.message, 'content')):
                    logging.error("Response lacks expected 'message.content' structure")
                    if expect_json:
                        return '{"error": "Invalid response format from AI"}'
                    else:
                        return "<html><body>Invalid response format from AI</body></html>"
                ai_content = response.message.content
                logging.info("Successfully received non-streaming response")

            # Clean up thinking tags
            ai_content = re.sub(r'<think>.*?</think>\s*', '', ai_content, flags=re.DOTALL)

            # Validate JSON if expected
            if expect_json:
                try:
                    parsed = extract_and_parse_json(ai_content)
                    return json.dumps(parsed)  # return clean stringified JSON
                except Exception as e:
                    logging.error(f"Failed to extract/parse JSON: {e}")
                    if attempt < max_retries - 1:
                        time.sleep(2 ** attempt)
                        continue
                    return '{"error": "Failed to get valid JSON after retries"}'
            
            # Add HTML wrapper if needed (non-JSON responses)
            if not expect_json and not ai_content.strip().startswith('<!DOCTYPE') and not ai_content.strip().startswith('<html') and not ai_content.strip().startswith('{'):
                ai_content = f"<html><body>{ai_content}</body></html>"

            return ai_content.strip()
            
        except Exception as e:
            logging.error(f"Error in get_ai_response (attempt {attempt + 1}/{max_retries}): {e}")
            
            if attempt < max_retries - 1:
                # Custom longer backoff: 30s → 60s → 120s (or adjust to start at 60s)
                wait_times = [120] # Option: fixed 60s then increase
                # Alternative for starting exactly at 60s and escalating: [60, 120, 240]
                
                wait_time = wait_times[attempt] if attempt < len(wait_times) else 120  # fallback to 120s if more retries
                
                logging.info(f"Retrying in {wait_time} seconds...")
                time.sleep(wait_time)
                continue
            else:
                raise

def search_hubspot_entities(entity_type, filters=None, properties=None, limit=100, sort=None, max_results=None):
    """
    Search HubSpot CRM entities with automatic pagination support.
    
    Args:
        entity_type (str): Type of entity - 'contacts', 'companies', or 'deals'
        filters (list): List of filter groups for search
        properties (list): List of properties to return
        limit (int): Results per page (default: 100, max: 200)
        sort (dict): Sorting configuration
        max_results (int): Maximum total results to fetch (None = fetch all)
    
    Returns:
        dict: Search results with 'results' containing list of all entities
    """
    try:
        endpoint = f"{HUBSPOT_BASE_URL}/crm/v3/objects/{entity_type}/search"
        
        headers = {
            "Authorization": f"Bearer {HUBSPOT_API_KEY}",
            "Content-Type": "application/json"
        }
        
        # Build base search payload
        payload = {
            "limit": min(limit, 200),
            "properties": properties or []
        }
        
        if filters:
            payload["filterGroups"] = [{"filters": filters}]
        
        if sort:
            payload["sorts"] = [sort]
        
        all_results = []
        after = None
        page = 1
        
        while True:
            # Add pagination parameter if available
            if after:
                payload["after"] = after
            
            logging.info(f"Fetching page {page} for {entity_type} (after={after})")
            
            # Make API request
            response = requests.post(endpoint, headers=headers, json=payload, timeout=30)
            response.raise_for_status()
            
            results = response.json()
            current_batch = results.get('results', [])
            all_results.extend(current_batch)
            
            logging.info(f"Page {page}: Retrieved {len(current_batch)} {entity_type} (Total: {len(all_results)})")
            
            # Check if we've reached max_results limit
            if max_results and len(all_results) >= max_results:
                all_results = all_results[:max_results]
                logging.info(f"Reached max_results limit of {max_results}")
                break
            
            # Check for pagination
            paging = results.get('paging', {})
            next_page = paging.get('next', {})
            after = next_page.get('after')
            
            # Break if no more pages
            if not after:
                logging.info(f"No more pages. Total {entity_type} fetched: {len(all_results)}")
                break
            
            page += 1
        
        return {"results": all_results, "total": len(all_results)}
        
    except requests.exceptions.RequestException as e:
        logging.error(f"HubSpot API error for {entity_type}: {str(e)}")
        if hasattr(e, 'response') and e.response is not None:
            logging.error(f"Response content: {e.response.text}")
        return {"results": [], "error": str(e)}
    except Exception as e:
        logging.error(f"Unexpected error searching {entity_type}: {str(e)}")
        return {"results": [], "error": str(e)}


def search_contacts(filters=None, properties=None, limit=200, max_results=None):
    """
    Search HubSpot contacts with pagination support.
    
    Args:
        filters (list): Filter criteria
        properties (list): Properties to return
        limit (int): Results per page (max 200)
        max_results (int): Maximum total results (None = fetch all)
    
    Returns:
        dict: Search results with all pages combined
    """
    default_properties = [
        'firstname', 'lastname', 'email', 'phone', 
        'jobtitle', 'address', 'city', 'state', 'zip', 'country',
        'createdate', 'lastmodifieddate', 'hs_object_id', 'hubspot_owner_id'
    ]
    
    return search_hubspot_entities(
        entity_type='contacts',
        filters=filters,
        properties=properties or default_properties,
        limit=limit,
        max_results=max_results
    )


def search_companies(filters=None, properties=None, limit=200, max_results=None):
    """
    Search HubSpot companies with pagination support.
    
    Args:
        filters (list): Filter criteria
        properties (list): Properties to return
        limit (int): Results per page (max 200)
        max_results (int): Maximum total results (None = fetch all)
    
    Returns:
        dict: Search results with all pages combined
    """
    default_properties = [
        'name', 'domain', 'phone',
        'address', 'city', 'state', 'zip', 'country',
        'industry', 'type', 'hs_object_id'
    ]
    
    return search_hubspot_entities(
        entity_type='companies',
        filters=filters,
        properties=properties or default_properties,
        limit=limit,
        max_results=max_results
    )


def search_deals(filters=None, properties=None, limit=200, max_results=None):
    """
    Search HubSpot deals with pagination support.
    
    Args:
        filters (list): Filter criteria
        properties (list): Properties to return
        limit (int): Results per page (max 200)
        max_results (int): Maximum total results (None = fetch all)
    
    Returns:
        dict: Search results with all pages combined
    """
    default_properties = [
        'dealname', 'amount', 'dealstage', 'pipeline', 'closedate',
        'createdate', 'hubspot_owner_id', 'hs_object_id'
    ]
    final_properties = list(set(default_properties + (properties or [])))
    logging.info(f"property is:{final_properties}")
    
    return search_hubspot_entities(
        entity_type='deals',
        filters=filters,
        properties=final_properties,
        limit=limit,
        max_results=max_results
    )

def export_to_file(data, export_format='excel', filename=None, export_dir='/appz/cache/exports', hyperlinks=None):
    try:
        os.makedirs(export_dir, exist_ok=True)
        if filename is None:
            filename = f"export_{uuid.uuid4()}"
        filename = os.path.splitext(filename)[0]

        filepath = os.path.join(export_dir, f"{filename}.xlsx")

        if isinstance(data, dict):
            data = list(data.values())[0] if len(data) == 1 else data

        if isinstance(data, (list, dict)):
            df = pd.DataFrame(data)
        elif isinstance(data, pd.DataFrame):
            df = data.copy()
        else:
            raise ValueError("Unsupported data type")
        
        # CRITICAL FIX: Remove timezone info from all datetime columns
        for col in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                df[col] = df[col].dt.tz_localize(None)

        # CRITICAL FIX: Sanitize all string columns to remove illegal Excel characters
        def sanitize_excel_value(val):
            """Remove illegal characters that openpyxl cannot write to Excel."""
            if pd.isna(val):
                return val
            if not isinstance(val, str):
                return val
            
            # Remove illegal XML characters (control characters except tab, newline, carriage return)
            # Excel XML spec doesn't allow: 0x00-0x08, 0x0B-0x0C, 0x0E-0x1F, 0x7F-0x9F
            illegal_chars = re.compile(
                r'[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F-\x9F]'
            )
            sanitized = illegal_chars.sub('', val)
            
            # Also replace other problematic characters with safe alternatives
            sanitized = sanitized.replace('\x0B', ' ')  # Vertical tab
            sanitized = sanitized.replace('\x0C', ' ')  # Form feed
            
            return sanitized

        # Apply sanitization to all columns
        for col in df.columns:
            if df[col].dtype == 'object':  # String columns
                df[col] = df[col].apply(sanitize_excel_value)

        # Keep helper columns for hyperlinks
        df_with_helpers = df.copy()

        # Define which columns to display (exclude helper columns)
        display_columns = [col for col in df.columns if not col.startswith('_')]
        df_display = df[display_columns]

        # Write to Excel
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df_display.to_excel(writer, sheet_name='Sheet1', index=False)
            worksheet = writer.sheets['Sheet1']

            # Apply formatting to data rows
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    col_letter = cell.column_letter
                    col_name = df_display.columns[cell.column - 1]  # 0-indexed

                    # Format Date columns (e.g., "Close Date", "Created Date", etc.)
                    if "date" in col_name.lower() or col_name in ["Close Date", "Expected Close Date", "Created At", "Last Modified"]:
                        if cell.value and isinstance(cell.value, datetime):
                            cell.number_format = 'MMM DD, YYYY'  # Changed to DD-MMM-YYYY

                    # Format Amount columns with $ currency
                    if col_name in ["Amount", "Deal Amount", "Total Value"]:
                        if cell.value:
                            try:
                                if isinstance(cell.value, str):
                                    # Clean and convert string amounts like "$1,234" or "1234"
                                    clean_val = re.sub(r'[^\d.]', '', str(cell.value))
                                    cell.value = float(clean_val) if clean_val else 0
                                else:
                                    cell.value = float(cell.value or 0)
                                cell.number_format = '$#,##0'  # US Dollar format with $ symbol
                            except:
                                pass

            # Apply hyperlinks if specified
            if hyperlinks and 'Sheet1' in hyperlinks:
                link_config = hyperlinks['Sheet1']
                for display_col_name, config in link_config.items():
                    if display_col_name not in df_display.columns:
                        logging.warning(f"Column {display_col_name} not found for hyperlink")
                        continue

                    url_col = config.get('url_column', '_hubspot_url')
                    if url_col not in df_with_helpers.columns:
                        logging.warning(f"URL column {url_col} not found")
                        continue

                    col_idx = df_display.columns.get_loc(display_col_name) + 1  # +1 for Excel
                    col_letter = get_column_letter(col_idx)

                    for row_idx in range(len(df_with_helpers)):
                        url = df_with_helpers.iloc[row_idx][url_col]
                        if pd.notna(url) and str(url).strip():
                            cell = worksheet[f"{col_letter}{row_idx + 2}"]  # +2: header + 0-index
                            cell.hyperlink = str(url)
                            cell.font = Font(color="0563C1", underline="single")
                            cell.style = "Hyperlink"

        logging.info(f"Exported with working hyperlinks: {filepath}")
        return {
            "success": True,
            "filepath": filepath,
            "filename": f"{filename}.xlsx",
            "format": "excel"
        }
    except Exception as e:
        logging.error(f"Export failed: {e}", exc_info=True)
        return {"success": False, "error": str(e)}
        
def branch_function(**kwargs):
    ti = kwargs['ti']
    unread_emails = ti.xcom_pull(task_ids="fetch_unread_emails", key="unread_emails")

    if not unread_emails:
        logging.info("No unread emails found, proceeding to no_email_found_task.")
        return "no_email_found_task"

    task_completion_emails = []
    other_emails = []
    logging.info(f"task completion email is :{task_completion_emails}")
    logging.info(f"other emails is :{other_emails}")
    
    for email in unread_emails:
        task_id = check_if_task_completion_reply(email)
        if task_id:
            # This is a task completion reply
            email["task_id"] = task_id
            task_completion_emails.append(email)
            logging.info(f"Identified task completion reply for task {task_id}")
        else:
            other_emails.append(email)
    
    # If we have task completion emails, route them separately
    if task_completion_emails:
        ti.xcom_push(key="task_completion_emails", value=task_completion_emails)
        logging.info(f"Routing {len(task_completion_emails)} task completion emails")
        return "trigger_task_completion"
    
    # For other emails, continue with existing routing logic
    if not other_emails:
        logging.info("No other emails to route after task completion separation")
        return "no_email_found_task"

    # Build email details for AI analysis with FULL chat history context
    email_details = []
    for idx, email in enumerate(unread_emails, 1):
        headers = email.get("headers", {})
        
        email_info = {
            "email_number": idx,
            "from": headers.get("From", "Unknown"),
            "subject": headers.get("Subject", "No Subject"),
            "latest_message": email.get("content", ""),
            "is_reply": email.get("is_reply", False)
        }
        email_details.append(email_info)
        logging.info(f"email details is: {email_details}")
    
    # Enhanced routing prompt with REPORT INTENT capability
    prompt = f"""You are an extremely strict, zero-tolerance email router for a HubSpot AI agent.
You have exactly 4 possible outputs. You must pick ONE and only.
Emails (with full thread context):
{email_details}
RETURN ONLY ONE OF THESE FOUR JSONS — NO TEXT BEFORE/AFTER:
{{"task_type": "search_dag"}}
{{"task_type": "continuation_dag"}}
{{"task_type": "report_dag"}}
{{"task_type": "no_action"}}
# RULES — MEMORIZE AND OBEY 100%
# ROUTE TO **search_dag** → FIRST-TIME ACTION (use only in these 8 cases):
    1. User wants to CREATE anything new → deal, contact, company, task, meeting, note, call log. Exclude the situation when the user is replying to a confirmation template.
    2. User asks for "360 view", "full picture", "deep dive", "what do we know about X", "research company"
    3. User pastes meeting notes/transcript and clearly expects them to be saved in HubSpot
    4. User gives a meeting minutes or a conversational prompt AND it's followed by a creation intent. exclude the situation when the user is confirming and adding changes to the confirmation mail.
    5. User explicitly says "summarize our history with Acme" (because this requires pulling engagements).Mainly used before the next meeting with the exiisting client
→ {{"task_type": "search_dag"}}

# ROUTE TO **continuation_dag** → USER IS REPLYING TO OUR CONFIRMATION EMAIL
    - This is detected when:
        • The thread already contains one of our confirmation emails and user is replying to that confirmation mail
        AND
        • Latest message contains any of these:
        - "yes", "proceed", "go ahead", "confirmed", "looks good", "perfect"
        - OR moving forward along with corrections/changes ("change amount to $200k", "add Sarah", "owner should be Mike", "close date Dec 31")
        - Is a casual comment regarding the mentioned meeting minutes in the thread ("Great call yesterday", ).
        - To create a followup task related to the meeting minutes in the thread.
        - Update any of the entities created in the converstaion history.(For example, "Change amount to $350k and add Sarah as contact")
        - # Casual comments definition:
            These are messages that do not request any action or information, such as:
            - "Will circle back next week"
            - "Sounds good"
            - "Looking forward to our meeting"
            - Comments for the client company or deal being discussed.
→ {{"task_type": "continuation_dag"}}

# ROUTE TO **report_dag** → ONLY when user explicitly asks for a report.
- Route to report_dag only when the user's prompt explicitly includes one of the following phrases: "generate report", "create report", "send report", "export", "quarterly report", or "deal report".
If none of these terms are present, do not trigger report_dag under any circumstances.”
- Do NOT route to report_dag when:
- The user requests details, information, summary, 360 view, overview, or insights without using the word “report”.
- Examples that must not trigger a report:
    - “Get the company details of Y”
    - “Show me the deal information.”
→ {{"task_type": "report_dag"}}

# ROUTE TO **no_action** → Only when user asks any query about hubspot data or just casually talks with the bot
Includes:
• Pure greetings: hi, hello, thanks, thank you, good morning, have a great day
• Questions about how the bot works: "How does this work?",:"What are your capabilities?"
• ALL data retrieval requests:
   - "Show me open deals"
   - "Any deals closing this month?"
   - "My tasks due today"
   - "Find contact John Doe"
   - "Is there a deal called Pipeline Booster?"
   - "What's the status of deal X?"
• Blank emails or just "?"
• You dont have the capability to act on casual comments or chit-chat.
→ {{"task_type": "no_action"}}

EXAMPLES — YOU MUST GET THESE 100% RIGHT

┌────────────────────────────────────────────────────────────────────┬──────────────────────┐
│ Latest user message                                                │ CORRECT ROUTE        │
├────────────────────────────────────────────────────────────────────┼──────────────────────┤
│ "Create a $300k deal for Nvidia closing Q4"                        │ search_dag           │
│ "Log today's call with Sarah from Stripe"                          │ search_dag           │
│ "Give me a 360 view of the Enterprise deal"                        │ search_dag           │
│ "Yes, proceed" (in thread with our confirmation)                   │ continuation_dag     │
│ "Change amount to $350k and add Sarah as contact"                  │ continuation_dag     │
│ "Looks good, just change close date to Dec 20"                     │ continuation_dag     │
│ "Generate quarterly sales report"                                  │ report_dag           │
│ "Send me a pipeline report"                                        │ report_dag           │
│ "Hi there!"                                                        │ no_action            │
│ "Thanks!"                                                          │ no_action            │
│ "Show all deals closing this month"                                │ no_action            │
│ "Any follow-ups due today?"                                        │ no_action            │
│ "Can you pull contact details for john@acme.com?"                  │ no_action            │
│ "THis was a great meeting, looking forward to our next steps"      │ continuation_dag     │
└────────────────────────────────────────────────────────────────────┴──────────────────────┘
Final instruction: If in doubt → route to **no_action**. Never guess creation**.

Return exactly one valid JSON. No reasoning field. No extra text.
"""
    logging.info(f"Sending routing prompt to AI with {len(email_details)} emails and conversation context")
    
    conversation_history_for_ai = []
    for email in unread_emails:
        chat_history = email.get("chat_history", [])
        for i in range(0, len(chat_history), 2):
            if i + 1 < len(chat_history):
                user_msg = chat_history[i]
                assistant_msg = chat_history[i + 1]
                if user_msg["role"] == "user" and assistant_msg["role"] == "assistant":
                    conversation_history_for_ai.append({
                        "prompt": user_msg["content"],
                        "response": assistant_msg["content"]
                    })

    response = get_ai_response(
        prompt=prompt,
        conversation_history=conversation_history_for_ai,
        expect_json=True
    )
    logging.info(f"AI routing response: {response}")   
    # Parse JSON response
    try:
        json_response = json.loads(response)
    except json.JSONDecodeError:
        json_response = extract_json_from_text(response)

    # Route based on AI decision
    if json_response and "task_type" in json_response:
        task_type = json_response["task_type"].lower()
        reasoning = json_response.get("reasoning", "No reasoning provided")
        logging.info(f"✓ AI DECISION: {task_type}, REASONING: {reasoning}")
        
        if "report" in task_type:
            ti.xcom_push(key="report_emails", value=unread_emails)
            logging.info(f"→ Routing {len(unread_emails)} emails to report_dag")
            return "trigger_report_dag"
            
        elif "continuation" in task_type:
            ti.xcom_push(key="reply_emails", value=unread_emails)
            logging.info(f"→ Routing {len(unread_emails)} emails to continuation_dag")
            return "trigger_continuation_dag"
            
        elif "search" in task_type:
            ti.xcom_push(key="new_emails", value=unread_emails)
            logging.info(f"→ Routing {len(unread_emails)} emails to search_dag")
            return "trigger_meeting_minutes"
        
        elif "no_action" in task_type:
            ti.xcom_push(key="no_action_emails", value=unread_emails)
            logging.info("→ No action needed for the emails")
            return "analyze_and_search_with_tools"


def extract_latest_reply(email_content):
    """
    Extract only the latest reply from an email, removing quoted conversation history.
    
    Args:
        email_content: Full email body text
    
    Returns:
        String containing only the latest message content
    """
    if not email_content:
        return ""
    
    # Split by common reply separators
    separators = [
        '\r\n\r\nOn ',  # Gmail style: "On Wed, Oct 22, 2025..."
        '\n\nOn ',
        '\r\n\r\n>',  # Quoted text starting with >
        '\n\n>',
        '\r\n\r\nFrom:',  # Outlook style
        '\n\nFrom:',
        '________________________________',  # Outlook horizontal line
    ]
    
    latest_content = email_content
    earliest_position = len(email_content)
    
    # Find the earliest occurrence of any separator
    for separator in separators:
        pos = email_content.find(separator)
        if pos != -1 and pos < earliest_position:
            earliest_position = pos
            latest_content = email_content[:pos]
    
    # Clean up the extracted content
    latest_content = latest_content.strip()
    
    # Remove leading ">" quoted lines if any remain
    lines = latest_content.split('\n')
    clean_lines = []
    for line in lines:
        stripped = line.lstrip()
        if not stripped.startswith('>'):
            clean_lines.append(line)
        else:
            break  # Stop at first quoted line
    
    return '\n'.join(clean_lines).strip()
def extract_json_from_text(text):
    """Extract JSON from text with markdown or other wrappers."""
    try:
        text = text.strip()
        text = re.sub(r'```json\s*', '', text)
        text = re.sub(r'```\s*', '', text)
        
        match = re.search(r'\{[^{}]*\}', text, re.DOTALL)
        if match:
            return json.loads(match.group())
        return None
    except Exception as e:
        logging.error(f"Error extracting JSON: {e}")
        return None

def trigger_meeting_minutes(**kwargs):
    ti = kwargs['ti']
    new_emails = ti.xcom_pull(task_ids="branch_task", key="new_emails") or []
    
    if not new_emails:
        logging.info("No new emails to process")
        return
    
    for email in new_emails:
        # Pass email with chat_history to search DAG
        trigger_conf = {
            "email_data": email,
            "chat_history": email.get("chat_history", []),
            "thread_history": email.get("thread_history", []),
            "thread_id": email.get("threadId", ""),  # ADD THIS
            "message_id": email.get("id", "")
        }
        
        task_id = f"trigger_search_{email['id'].replace('-', '_')}"
        trigger_task = TriggerDagRunOperator(
            task_id=task_id,
            trigger_dag_id="hubspot_search_entities",
            conf=trigger_conf,
        )
        trigger_task.execute(context=kwargs)
    
    logging.info(f"Triggered search DAG for {len(new_emails)} emails")

def trigger_continuation_dag(**kwargs):
    ti = kwargs['ti']
    reply_emails = ti.xcom_pull(task_ids="branch_task", key="reply_emails") or []
    
    if not reply_emails:
        logging.info("No reply emails to process")
        return

    for email in reply_emails:
        # Pass email with full chat_history to continuation DAG
        trigger_conf = {
            "email_data": email,
            "chat_history": email.get("chat_history", []),
            "thread_history": email.get("thread_history", []),
            "thread_id": email.get("threadId"),
            "thread_id": email.get("threadId", ""),  # ADD THIS
            "message_id": email.get("id", "")
        }
        
        task_id = f"trigger_continuation_{email['id'].replace('-', '_')}"
        trigger_task = TriggerDagRunOperator(
            task_id=task_id,
            trigger_dag_id="hubspot_create_objects",
            conf=trigger_conf,
        )
        trigger_task.execute(context=kwargs)
        
        logging.info(f"✓ Triggered continuation DAG for thread {email.get('threadId')}")

    logging.info(f"Triggered continuation for {len(reply_emails)} emails")

def generate_spelling_variants_for_prompt(user_content, chat_history=None):
    """
    Generate potential spelling variants for contact/company/deal names.
    Returns (inject_text: str for prompt, variants_dict: dict)
    """
    # Build recent context for better detection
    context_lines = []
    if chat_history:
        for msg in chat_history[-4:]:  # Last 4 messages
            role = msg.get("role", "user")
            content = msg.get("content", "")
            context_lines.append(f"{role.upper()}: {content}")
    context_lines.append(f"USER: {user_content}")
    full_context = "\n\n".join(context_lines)

    variant_prompt = f"""You are a spelling variant assistant for a HubSpot email assistant.

CRITICAL RULES:
1. ONLY suggest variants if the name looks OBVIOUSLY misspelled (missing letters, wrong letters)
2. If a name looks normal, return empty lists - DON'T suggest alternatives
3. Only suggest variants for garbled text like "Snidhu", "Prya", "Jhon"

Examples of when to suggest variants:
- "Snidhu" → ["snidhu", "sindhu"] (one letter wrong)
- "Micrsoft" → ["micrsoft", "microsoft"] (missing letter)

Examples of when NOT to suggest variants:
- "Sindhu" → [] (already correct)
- "Priya" → [] (common name)
- "Microsoft" → [] (known company)

Extract any contact names, company names, or deal names from the latest user message that might have typos.
For each potentially misspelled name, give 3–5 plausible lowercase variants (include original as first).

Rules:
- Always include the original spelling first.
- Only suggest variants if it looks like a typo.
- Max 5 variants per name.
- Use lowercase only.
- If no likely typos, return empty lists.

CONVERSATION CONTEXT:
{full_context}

Return ONLY valid JSON:
{{
    "potential_variants": {{
        "contacts": [{{"original": "Neah", "variants": ["neah", "neha", "neeha", "nehha"]}}],
        "companies": [{{"original": "Microsft", "variants": ["microsft", "microsoft", "micrsoft"]}}],
        "deals": []
    }},
    "has_potential_typos": true|false
}}

If names look correct, return:
{{
    "potential_variants": {{"contacts": [], "companies": [], "deals": []}},
    "has_potential_typos": false
}}
"""

    try:
        response = get_ai_response(variant_prompt, expect_json=True)
        data = json.loads(response.strip())
        variants = data.get("potential_variants", {})
        has_typos = data.get("has_potential_typos", False)

        if not has_typos or not any(variants.values()):
            return "", {}

        inject_text = """
IMPORTANT SPELLING VARIANTS DETECTED:
The user may have misspelled names. Use the most plausible/correct spelling when interpreting the request, building filters, or report titles:

"""
        if variants.get("contacts"):
            inject_text += f"CONTACTS: {json.dumps(variants['contacts'], indent=2)}\n"
        if variants.get("companies"):
            inject_text += f"COMPANIES: {json.dumps(variants['companies'], indent=2)}\n"
        if variants.get("deals"):
            inject_text += f"DEALS: {json.dumps(variants['deals'], indent=2)}\n"

        inject_text += "\nPrefer the correct spelling (usually the 2nd item) in searches and titles.\n"

        logging.info(f"Spelling variants generated: {variants}")
        return inject_text, variants

    except Exception as e:
        logging.warning(f"Spelling variant generation failed: {e}")
        return "", {}

def analyze_and_search_with_tools(**kwargs):
    """LLM analyzes email and performs search if needed. Handles ambiguity and cross-entity queries."""
    import re  # Import at the beginning of the function
    
    ti = kwargs['ti']
    unread_emails = ti.xcom_pull(task_ids="branch_task", key="no_action_emails") or []
    
    if not unread_emails:
        logging.info("No emails requiring friendly response")
        return
    
    service = authenticate_gmail()
    if not service:
        logging.error("Gmail authentication failed, cannot send responses")
        return

    processed_emails = []

    for email in unread_emails:
        email_id = email.get("id", "unknown")
        content = email.get("content", "").strip()
        headers = email.get("headers", {})
        sender_email = headers.get("From", "")
        thread_history = email.get("thread_history", [])
        chat_history = email.get("chat_history", [])  # ← ADD THIS LINE

        sender_name = "there"
        name_match = re.search(r'^([^<]+)', sender_email)
        if name_match:
            sender_name = name_match.group(1).strip().split()[0] if name_match.group(1).strip() else "there"

        # === ADD SPELLING VARIANT DETECTION HERE ===
        spelling_inject_text, spelling_variants = generate_spelling_variants_for_prompt(
            user_content=content,
            chat_history=chat_history  # ← Pass full history
        )
        if spelling_variants:
            logging.info(f"SPELLING VARIANTS DETECTED: {json.dumps(spelling_variants, indent=2)}")
        else:
            logging.info("No spelling variants detected - names appear correct")

        # Check if spelling correction was applied
        was_typo_corrected = False
        closest_match_name = None

        if spelling_variants:
            # Check if any variants were used in the search
            for entity_type in ['contacts', 'companies', 'deals']:
                variants_list = spelling_variants.get(entity_type, [])
                for variant_group in variants_list:
                    original = variant_group.get('original', '').lower()
                    variants = variant_group.get('variants', [])
                    
                    # If we have variants, the second one is usually the corrected version
                    if len(variants) > 1 and original != variants[1]:
                        was_typo_corrected = True
                        closest_match_name = variants[1].title()  # Capitalize properly
                        logging.info(f"✓ Detected typo correction: '{original}' → '{closest_match_name}'")
                        break
                if was_typo_corrected:
                    break

        try:
            # CHECK IF CLARIFICATION WAS ALREADY SENT IN THIS THREAD
            clarification_context = ""
            clarification_already_sent = False
            original_query_type = None
            target_entity = None
            original_query = None
            
            for msg in reversed(thread_history):  # Latest bot message first
                if msg.get("from_bot", False):
                    bot_content = msg.get("content", "")
                    # Look for metadata markers in the HTML content
                    if ("Multiple Matches Found" in bot_content or 
                        "please clarify which one" in bot_content.lower() or
                        "Attached File" in bot_content):
                        clarification_already_sent = True
                        
                        # Try to extract metadata from HTML comments (if stored)
                        query_type_match = re.search(r'<!--\s*QUERY_TYPE:\s*(\S+)\s*-->', bot_content)
                        target_entity_match = re.search(r'<!--\s*TARGET_ENTITY:\s*(\S+)\s*-->', bot_content)
                        original_query_match = re.search(r'<!--\s*ORIGINAL_QUERY:\s*(.+?)\s*-->', bot_content)
                        
                        if query_type_match:
                            original_query_type = query_type_match.group(1)
                        if target_entity_match:
                            target_entity = target_entity_match.group(1)
                        if original_query_match:
                            original_query = original_query_match.group(1)
                        
                        clarification_context = f"""
IMPORTANT: A clarification request was previously sent to the user.
Original user request: "{original_query if original_query else 'get associated data'}"
Query type: {original_query_type if original_query_type else 'unknown'}
Target entity: {target_entity if target_entity else 'unknown'}
User is now replying with: "{content}"

INSTRUCTIONS FOR CLARIFICATION RESPONSE:
1. Parse the user's response to identify which specific entity they're selecting:
   - If they mention a row number (e.g., "row 1", "first one"), use that
   - If they provide an email address, match by email
   - If they provide a name with company, match by that combination
   
2. Extract the entity ID from the clarification
   
3. CRITICAL: If the original query type contains "for" (like "company_for_contact", "deals_for_contact"):
   This means they want ASSOCIATED data, not the entity itself.
   After identifying the entity:
   - Use the entity ID to search for the TARGET entity
   - For "company_for_contact": use contact_id to search companies
   - For "deals_for_contact": use contact_id to search deals
   - For "contacts_for_company": use company_id to search contacts
   
4. Return the FINAL results (the target entity data) with proper fields
"""
                        break
            
            # MAIN PROMPT
            prompt = f"""You are a friendly HubSpot email assistant with the following capabilities:

CRITICAL RULE: ALWAYS fetch ALL matching results up to 200 maximum. NEVER use pagination or limits of 10.
When searching, use max_results=200 parameter to get everything at once.

- Answer generic HubSpot questions
- Handle greetings and casual conversation
- Search HubSpot data (contacts, companies, deals, tasks)
- Handle associated data queries (e.g., "deals for a contact", "company for a contact")
- consider spelling variants when searching: {spelling_inject_text}

User message: "{content}"
{clarification_context}

═══════════════════════════════════════════════════════════════════
REQUIRED OUTPUT FIELDS BY ENTITY TYPE:
═══════════════════════════════════════════════════════════════════

**CONTACTS:** Always return these fields in this order:
- Contact ID
- Firstname
- Lastname
- Email
- Phone
- Job Title
- Contact Owner Name
- Last Modified Date

**COMPANIES:** Always return these fields in this order:
- Company Id
- Company Name
- Domain

**DEALS:** Always return these fields in this order:
- Deal ID
- Deal Name
- Deal Stage (use human-readable name, e.g., "APPOINTMENT SCHEDULED" not "appointmentscheduled")
- Deal Owner
- Deal Amount
- Expected Close Date
- Associated Company
- Associated Contacts

**TASKS:** Always return these fields in this order:
- Task ID
- Task Name
- Status
- Priority
- Due Date
- Owner

═══════════════════════════════════════════════════════════════════
CRITICAL WORKFLOW:
═══════════════════════════════════════════════════════════════════
IMPORTANT:
**Pagination Handling:**
**PAGINATION - CRITICAL OVERRIDE:**
- ALWAYS set max_results=200 in ALL search calls
- NEVER paginate or limit to 10 results
- Example: search_contacts(filters=[...], max_results=200)
- Example: search_deals(filters=[...], max_results=200)
- If you see ANY pagination code, DELETE IT and use max_results=200

**1. IDENTIFY QUERY TYPE:**
   - Direct entity query: "show me all contacts", "find company X"
   - Associated data query: "deals for contact X", "company for contact Y", "contacts in company Z"
   - Cross-entity query: "tasks for deal X", "deals for company Y"

**2. FOR DIRECT QUERIES:**
   - Contact query → call `search_contacts`
   - Company query → call `search_companies`
   - Deal query → call `search_deals`
   - Task query → first get owner_id via `get_all_owners`, then call `search_tasks`

**3. FOR ASSOCIATED DATA QUERIES (CRITICAL):**
   Step 1: Search for PRIMARY entity first
   Step 2: If multiple matches found → Mark as ambiguous and return:
      {{
        "is_ambiguous": true,
        "requires_clarification": true,
        "ambiguous_entity": "contacts|companies|deals",
        "results": [...all matches with ALL required fields...],
        "result_count": <count>,
        "search_term": "what user searched for",
        "metadata": {{
          "query_type": "company_for_contact|deals_for_contact|contacts_for_company|deals_for_company|etc",
          "original_query": "{content}",
          "target_entity": "companies|deals|contacts"
        }}
      }}
   Step 3: After clarification received:
      - Parse user's clarification response
      - Identify the specific entity they selected
      - Extract the entity ID
      - NOW search for the TARGET entity using this ID
      - Return final results with proper fields

**4. HANDLING CLARIFICATION RESPONSES (CRITICAL):**
   When clarification_already_sent is True:
   
   a) PARSE USER'S CLARIFICATION:
      - Extract row number: "row 1", "first one", "the second", "number 3"
      - Extract email: look for @domain.com pattern
      - Extract name + company: "John at Acme", "Sarah Smith from Tech Corp"
   
   b) MATCH TO ORIGINAL RESULTS:
      - Row number: use results[row_number]
      - Email: find result where Email field matches exactly
      - Name + company: find best match combining name fields and company
   
   c) GET THE ENTITY ID:
      - Extract the ID field from the matched result
      - For contacts: use "Contact ID" or "id" field
      - For companies: use "Company Id" or "id" field
      - For deals: use "Deal ID" or "id" field
   
   d) CHECK QUERY TYPE:
      - If query_type == "company_for_contact":
        * Use the contact ID you just found
        * Call search_companies with filter: {{"propertyName": "associations.contact", "operator": "CONTAINS_TOKEN", "value": "<contact_id>"}}
        * Return company results with proper Company fields
      
      - If query_type == "deals_for_contact":
        * Use the contact ID you just found
        * Call search_deals with filter: {{"propertyName": "associations.contact", "operator": "CONTAINS_TOKEN", "value": "<contact_id>"}}
        * Return deal results with proper Deal fields
      
      - If query_type == "contacts_for_company":
        * Use the company ID you just found
        * Call search_contacts with filter: {{"propertyName": "associatedcompanyid", "operator": "EQ", "value": "<company_id>"}}
        * Return contact results with proper Contact fields
      
      - If query_type == "direct_search" (no association):
        * Return the selected entity directly - no additional search needed
        * Just return that one contact/company/deal they selected
   
   e) RETURN FORMAT:
      {{
        "clarification_resolved": true,
        "needs_search": true,
        "results": [...TARGET entity results with ALL required fields...],
        "result_count": <count>,
        "entity": "<target_entity_type>",
        "clarified_entity_id": "<the ID they selected>",
        "selected_from_original": "<original entity info for logging>"
      }}
   
   🚨 CRITICAL: After clarification, return ONLY the TARGET entity data they asked for.
   If they wanted "deals for Priya", don't return Priya's contact info - return her DEALS.

═══════════════════════════════════════════════════════════════════
SEARCH GUIDELINES:
═══════════════════════════════════════════════════════════════════

- Search by partial names (first or last name only is fine)
- Fetch ALL results up to 200 max (no pagination in email)
- Keep data single-line (no wrapping for phone/email)
- Phone format: +91 98765 43210
- Use "N/A" for empty fields (never blank)
- Dates: YYYY-MM-DD format only

═══════════════════════════════════════════════════════════════════
DATE HANDLING FOR DEALS:
═══════════════════════════════════════════════════════════════════

- Parse natural language: "this month", "next week", "Q4 2024"
- Convert to date ranges: start_date (GTE) and end_date (LTE)
- "This month end" = GTE:today, LTE:last_day_of_current_month
- Always include today when relevant
- Sort deals by Expected Close Date (ascending)

═══════════════════════════════════════════════════════════════════
AMBIGUITY DETECTION (CRITICAL):
═══════════════════════════════════════════════════════════════════

Mark as ambiguous ONLY when ALL conditions met:
1. User wants ASSOCIATED data (not a direct list)
2. Primary entity has MULTIPLE matches
3. User hasn't provided unique identifier (no email, no company context)

Examples:
✓ AMBIGUOUS: "Get deals for Priya" + finds 3 Priyas + no email given
✓ AMBIGUOUS: "Company for John Smith" + finds 5 John Smiths + no unique info
✗ NOT ambiguous: "Show all contacts named Sarah" (they want the list)
✗ NOT ambiguous: "Deals for priya@company.com" (unique identifier given)
✗ NOT ambiguous: "Companies in Tech sector" (direct company search)

═══════════════════════════════════════════════════════════════════
MANDATORY OUTPUT FORMAT:
═══════════════════════════════════════════════════════════════════

**If no search needed (greetings/thanks):**
{{"needs_search": false, "response_type": "casual"}}

**If search performed successfully:**
{{
  "needs_search": true,
  "results": [...exact results with required fields in json format...],
  "result_count": <number>,
  "entity": "contacts|companies|deals|tasks"
}}

**If ambiguous (clarification needed):**
{{
  "needs_search": true,
  "is_ambiguous": true,
  "requires_clarification": true,
  "ambiguous_entity": "contacts|companies|deals",
  "results": [...all matching entities with ALL required fields in json format...],
  "result_count": <number>,
  "search_term": "what user searched for",
  "metadata": {{
    "query_type": "company_for_contact|deals_for_contact|direct_search|etc",
    "original_query": "{content}",
    "target_entity": "companies|deals|contacts|none"
  }}
}}

**If clarification received and resolved:**
{{
  "needs_search": true,
  "clarification_resolved": true,
  "results": [...final target entity results with required fields in json format...],
  "result_count": <number>,
  "entity": "contacts|companies|deals|tasks",
  "clarified_entity_id": "12345"
}}

Always return valid JSON. No extra text or tables.
"""

            # Get AI response
            ai_response = get_ai_response(prompt=prompt, expect_json=True)
            logging.info(f"AI response for {email_id}: {ai_response}")

            # Extract JSON if wrapped in <think> or other tags, or has extra text
            if isinstance(ai_response, str):
                # Remove any <think>...</think> blocks and leading/trailing whitespace
                cleaned = re.sub(r'<think>.*?</think>', '', ai_response, flags=re.DOTALL)
                cleaned = cleaned.strip()
                
                # Find the first { or [ and extract valid JSON from there
                json_start = cleaned.find('{')
                json_end = cleaned.rfind('}') + 1
                if json_start == -1 or json_end == 0:
                    raise ValueError("No valid JSON found in AI response")
                
                json_str = cleaned[json_start:json_end]
                
                try:
                    analysis = json.loads(json_str)
                except json.JSONDecodeError as e:
                    logging.error(f"Failed to parse extracted JSON: {json_str}")
                    raise e
            else:
                analysis = ai_response

            # HANDLE AMBIGUITY - SEND CLARIFICATION WITH METADATA EMBEDDED IN HTML
            if analysis.get("requires_clarification", False) and analysis.get("is_ambiguous", False):
                logging.info(f"⚠️ AMBIGUITY DETECTED for {email_id}")
                
                results = analysis.get("results", [])
                entity = analysis.get("ambiguous_entity", "records")
                search_term = analysis.get("search_term", "the search term")
                total_count = analysis.get("result_count", len(results))
                metadata = analysis.get("metadata", {})
                
                query_type = metadata.get("query_type", "direct_search")
                target_entity = metadata.get("target_entity", "none")
                original_query = metadata.get("original_query", content)
                
                # Send clarification based on count
                if total_count > 10:
                    logging.info(f"Count > 10 ({total_count}), exporting to Excel")
                    
                    df = pd.DataFrame(results)
                    export_result = export_to_file(
                        data=df,
                        export_format='excel',
                        filename=f"clarification_{entity}_{email_id[:8]}",
                        export_dir='/appz/cache/exports'
                    )
                    
                    if export_result.get("success"):
                        excel_path = export_result["filepath"]
                        excel_filename = export_result["filename"]
                        
                        clarification_html = build_clarification_email_with_excel(
                            sender_name=sender_name,
                            entity=entity,
                            search_term=search_term,
                            total_count=total_count,
                            query_type=query_type,
                            target_entity=target_entity,
                            original_query=original_query,
                            closest_match_name=closest_match_name,  # ADD THIS
                            was_typo=was_typo_corrected  # ADD THIS
                        )
                        
                        send_email_with_attachment(
                            service=service,
                            email=email,
                            html_content=clarification_html,
                            attachment_path=excel_path,
                            attachment_filename=excel_filename
                        )
                        
                        logging.info(f"✓ Sent clarification with Excel for {email_id}")
                    else:
                        logging.error(f"Excel export failed: {export_result.get('error')}")
                        clarification_html = build_clarification_email_html(
                            sender_name=sender_name,
                            total_count=total_count,
                            entity=entity,
                            search_term=search_term,
                            query_type=query_type,
                            target_entity=target_entity,
                            original_query=original_query,
                            closest_match_name=closest_match_name,  # ADD THIS
                            was_typo=was_typo_corrected  # ADD THIS
                        )
                        send_email_reply(service, email, clarification_html)
                else:
                    # <=10 results: HTML table
                    clarification_html = build_clarification_email_html(
                        sender_name=sender_name,
                        results=results,
                        entity=entity,
                        search_term=search_term,
                        total_count=total_count,
                        query_type=query_type,
                        target_entity=target_entity,
                        original_query=original_query
                    )
                    send_email_reply(service, email, clarification_html)
                    logging.info(f"✓ Sent clarification HTML for {email_id}")
                
                mark_message_as_read(service, email_id)
                continue

            # HANDLE CLARIFICATION RESOLUTION
            # HANDLE CLARIFICATION RESOLUTION
            if analysis.get("clarification_resolved", False):
                logging.info(f"✓ Clarification resolved for {email_id}")
                
                # Results should contain the FINAL TARGET entity data
                results = analysis.get("results", [])
                entity = analysis.get("entity", "records")
                result_count = analysis.get("result_count", len(results))
                clarified_entity_id = analysis.get("clarified_entity_id", "unknown")
                
                logging.info(f"Final results after clarification: {result_count} {entity} (clarified entity ID: {clarified_entity_id})")
                
                # ⭐ CRITICAL FIX: Don't send email here, let generate_final_response_or_trigger_report handle it
                # Just add to processed_emails list so Task 2 can format and send it properly
                processed_emails.append({
                    "email": email,
                    "sender_name": sender_name,
                    "analysis": analysis  # This contains the resolved results
                })
                
                # DON'T mark as read yet - Task 2 will do it after sending
                logging.info(f"✓ Passed clarification results to Task 2 for {email_id}")
                continue

            # NORMAL PROCESSING - no ambiguity
            processed_emails.append({
                "email": email,
                "sender_name": sender_name,
                "analysis": analysis
            })

        except Exception as e:
            logging.error(f"Task 1 FAILED for {email_id}: {e}", exc_info=True)
            raise 
        finally:
                mark_message_as_read(service, email_id)

    if processed_emails:
        ti.xcom_push(key="llm_search_results", value=processed_emails)


# ═══════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS - UPDATE THESE IN YOUR CODEBASE
# ═══════════════════════════════════════════════════════════════════

def build_clarification_email_html(sender_name, results, entity, search_term, total_count, query_type="direct_search", target_entity="none", original_query="",closest_match_name=None, was_typo=False):
    """Build HTML email for clarification with metadata embedded as HTML comments."""
    
    entity_display = entity.replace("_", " ").title()
    
    # Build table rows
    if entity == "contacts":
        header_row = "<th>Row</th><th>Contact ID</th><th>Name</th><th>Email</th><th>Phone</th>"
        table_rows = ""
        for idx, result in enumerate(results[:10], 1):
            name = f"{result.get('Firstname', '')} {result.get('Lastname', '')}".strip() or "N/A"
            table_rows += f"""
            <tr>
                <td>{idx}</td>
                <td>{result.get('Contact ID', result.get('id', 'N/A'))}</td>
                <td>{name}</td>
                <td>{result.get('Email', 'N/A')}</td>
                <td>{result.get('Phone', 'N/A')}</td>
            </tr>"""
    
    elif entity == "companies":
        header_row = "<th>Row</th><th>Company ID</th><th>Company Name</th><th>Domain</th>"
        table_rows = ""
        for idx, result in enumerate(results[:10], 1):
            table_rows += f"""
            <tr>
                <td>{idx}</td>
                <td>{result.get('Company Id', result.get('id', 'N/A'))}</td>
                <td>{result.get('Company Name', result.get('name', 'N/A'))}</td>
                <td>{result.get('Domain', result.get('domain', 'N/A'))}</td>
            </tr>"""
    
    elif entity == "deals":
        header_row = "<th>Row</th><th>Deal ID</th><th>Deal Name</th><th>Amount</th><th>Stage</th>"
        table_rows = ""
        for idx, result in enumerate(results[:10], 1):
            table_rows += f"""
            <tr>
                <td>{idx}</td>
                <td>{result.get('Deal ID', result.get('id', 'N/A'))}</td>
                <td>{result.get('Deal Name', result.get('dealname', 'N/A'))}</td>
                <td>{result.get('Deal Amount', result.get('amount', 'N/A'))}</td>
                <td>{result.get('Deal Stage', result.get('dealstage', 'N/A'))}</td>
            </tr>"""
    
    # Build query type message
    query_type_msg = ""
    if "for" in query_type:
        target_display = target_entity.replace("_", " ").title() if target_entity != "none" else "associated data"
        query_type_msg = f"<p><strong>Note:</strong> Once you clarify which {entity_display.lower()} you're referring to, I'll fetch the associated <strong>{target_display}</strong> for you.</p>"
    
    if was_typo:
        closest_text = f" <strong>{closest_match_name}</strong>" if closest_match_name else " a close match"
        intro_message = f"""
        <p>I couldn't find an exact match for "<strong>{search_term}</strong>", 
        but the closest match appears to be{closest_text}.</p>
        <p>Here are the top results — did you mean one of these?</p>
        """
    else:
        intro_message = f"""
        <p>I found <strong>{total_count} {entity_display}</strong> matching "<strong>{search_term}</strong>".</p>
        """
    
    # EMBED METADATA AS HTML COMMENTS (invisible to user, readable by code)
    metadata_comments = f"""
<!-- QUERY_TYPE: {query_type} -->
<!-- TARGET_ENTITY: {target_entity} -->
<!-- ORIGINAL_QUERY: {original_query} -->
"""
    
    html = f"""
<html>
<head>
    <style>
        body {{ font-family: Arial, sans-serif; padding: 20px; max-width: 900px; margin: 0 auto; }}
        .content {{ margin: 20px 0; }}
        table {{ border-collapse: collapse; width: 100%; margin: 20px 0; font-size: 0.9em; }}
        th {{ background-color: #f0f0f0; padding: 10px; text-align: left; border: 1px solid #ddd; }}
        td {{ padding: 10px; border: 1px solid #ddd; }}
        tr:nth-child(even) {{ background-color: #f9f9f9; }}
        .footer {{ margin-top: 20px; font-size: 0.9em; color: #666; }}
        .note {{ background-color: #fff3cd; padding: 10px; border-left: 4px solid #ffc107; margin: 15px 0; }}
    </style>
</head>
<body>
{metadata_comments}
    <div class="content">
        <p>Hello {sender_name},</p>
        {intro_message}
        {query_type_msg}
        
        <p>Please reply with one of the following to help me identify the specific {entity_display.lower()}:</p>
        <ul>
            <li><strong>Row number</strong> from the table below (e.g., "row 1" or "first one")</li>
            <li><strong>Email address</strong> (e.g., "priya.desai@northpeaksys.com")</li>
            <li><strong>Full name with company</strong> (e.g., "Priya Desai at NorthPeak")</li>
        </ul>
        
        <table>
            <thead>
                <tr>{header_row}</tr>
            </thead>
            <tbody>
                {table_rows}
            </tbody>
        </table>
        
        {'<p><em>Showing the matches. Reply with identifying information to proceed.</em></p>' if total_count > 10 else ''}
    </div>
    
    <div class="signature">
        <p>Best regards,<br>
        The HubSpot Assistant Team<br>
        <a href="http://lowtouch.ai" class="company">lowtouch.ai</a></p>
    </div>
</body>
</html>
"""
    return html

def build_clarification_email_with_excel(sender_name, total_count, entity, search_term,query_type="direct_search", target_entity="none", original_query="",closest_match_name=None, was_typo=False):
    """Build HTML email for clarification when sending Excel attachment."""
    
    entity_display = entity.replace("_", " ").title()
    
    query_type_msg = ""
    if "for" in query_type:
        target_display = target_entity.replace("_", " ").title() if target_entity != "none" else "associated data"
        query_type_msg = f"""
        <div class="note">
            <strong>Note:</strong> Once you clarify which {entity_display.lower()} you're referring to, 
            I'll fetch the associated <strong>{target_display}</strong> for you.
        </div>
        """
    if was_typo:
        closest_text = f" <strong>{closest_match_name}</strong>" if closest_match_name else " a close match"
        intro_message = f"""
        <p>I couldn't find an exact match for "<strong>{search_term}</strong>", 
        but the closest match appears to be{closest_text}.</p>
        <p>Please review the attached Excel file — did you mean one of these?</p>
        """
    else:
        intro_message = f"""
        <p>I found <strong>{total_count} {entity_display}</strong> matching "<strong>{search_term}</strong>".</p>
        """
    
    # EMBED METADATA AS HTML COMMENTS
    metadata_comments = f"""
<!-- QUERY_TYPE: {query_type} -->
<!-- TARGET_ENTITY: {target_entity} -->
<!-- ORIGINAL_QUERY: {original_query} -->
"""
    
    html = f"""
<html>
<head>
    <style>
        body {{ font-family: Arial, sans-serif; padding: 20px; max-width: 800px; margin: 0 auto; }}
        .content {{ margin: 20px 0; }}
        .note {{padding: 10px;margin: 15px 0; }}
        .footer {{ margin-top: 20px; font-size: 0.9em; color: #666; }}
    </style>
</head>
<body>
{metadata_comments}
    
    <div class="content">
        <p>Hello {sender_name},</p>
        
        {intro_message}
        {query_type_msg}
        
        <p>Please review the attached Excel file and reply with one of the following:</p>
        <ul>
            <li><strong>Row number</strong> (e.g., "row 15")</li>
            <li><strong>Email address</strong> (e.g., "contact@company.com")</li>
            <li><strong>Full name with company</strong> (e.g., "John Smith at Acme Corp")</li>
        </ul>
    </div>
    
    <div class="signature">
        <p>Best regards,<br>
        The HubSpot Assistant Team<br>
        <a href="http://lowtouch.ai" class="company">lowtouch.ai</a></p>
    </div>
</body>
</html>
"""
    return html

def build_no_results_html(sender_name, entity):
    """Build HTML for when no results are found"""
    entity_display = entity.replace("_", " ").title()
    
    return f"""
<html>
<head>
    <style>
        body {{
            font-family: Arial, sans-serif;
            line-height: 1.6;
            color: #333;
            max-width: 600px;
            margin: 0 auto;
            padding: 20px;
        }}
        .greeting {{ margin-bottom: 15px; }}
        .message {{ margin: 15px 0; }}
        .closing {{ margin-top: 15px; }}
        .signature {{ margin-top: 15px; font-weight: bold; }}
        .company {{ color: #666; font-size: 0.9em; }}
    </style>
</head>
<body>
    <div class="greeting">
        <p>Hello {sender_name},</p>
    </div>
    
    <div class="message">
        <p>I didn't find any {entity_display} in HubSpot that match the criteria you shared.</p>
    </div>
    
    <div class="closing">
        <p>If you need additional information, please don't hesitate to ask.</p>
    </div>
    
    <div class="signature">
        <p>Best regards,<br>
        The HubSpot Assistant Team<br>
        <a href="http://lowtouch.ai" class="company">lowtouch.ai</a></p>
    </div>
</body>
</html>
"""

def build_single_result_html(sender_name, result, entity):
    """Build HTML for single result (inline format)"""
    entity_display = entity.replace("_", " ").title()
    
    # Build details list based on entity type
    details_html = ""
    
    if entity == "contacts":
        contact_id = result.get('Contact ID', result.get('id', 'N/A'))
        firstname = result.get('Firstname', result.get('firstname', ''))
        lastname = result.get('Lastname', result.get('lastname', ''))
        name = f"{firstname} {lastname}".strip() or "N/A"
        email = result.get('Email', result.get('email', 'N/A'))
        phone = result.get('Phone', result.get('phone', 'N/A'))
        job_title = result.get('Job Title', result.get('jobtitle', 'N/A'))
        owner = result.get('Contact Owner Name', result.get('hubspot_owner_name', 'N/A'))
        
        details_html = f"""
        <ul>
            <li><strong>Contact ID:</strong> {contact_id}</li>
            <li><strong>Name:</strong> {name}</li>
            <li><strong>Email:</strong> {email}</li>
            <li><strong>Phone:</strong> {phone}</li>
            <li><strong>Job Title:</strong> {job_title}</li>
            <li><strong>Contact Owner:</strong> {owner}</li>
        </ul>
        """
    
    elif entity == "companies":
        company_id = result.get('Company Id', result.get('id', 'N/A'))
        company_name = result.get('Company Name', result.get('name', 'N/A'))
        domain = result.get('Domain', result.get('domain', 'N/A'))
        phone = result.get('Phone', result.get('phone', 'N/A'))
        industry = result.get('Industry', result.get('industry', 'N/A'))
        
        details_html = f"""
        <ul>
            <li><strong>Company ID:</strong> {company_id}</li>
            <li><strong>Company Name:</strong> {company_name}</li>
            <li><strong>Domain:</strong> {domain}</li>
            <li><strong>Phone:</strong> {phone}</li>
            <li><strong>Industry:</strong> {industry}</li>
        </ul>
        """
    
    elif entity == "deals":
        deal_id = result.get('Deal ID', result.get('id', 'N/A'))
        deal_name = result.get('Deal Name', result.get('dealname', 'N/A'))
        amount = result.get('Deal Amount', result.get('amount', 'N/A'))
        stage = result.get('Deal Stage', result.get('dealstage', 'N/A'))
        owner = result.get('Deal Owner', result.get('hubspot_owner_name', 'N/A'))
        close_date = result.get('Expected Close Date', result.get('closedate', 'N/A'))
        
        details_html = f"""
        <ul>
            <li><strong>Deal ID:</strong> {deal_id}</li>
            <li><strong>Deal Name:</strong> {deal_name}</li>
            <li><strong>Amount:</strong> {amount}</li>
            <li><strong>Deal Stage:</strong> {stage}</li>
            <li><strong>Close Date:</strong> {close_date}</li>
            <li><strong>Deal Owner:</strong> {owner}</li>
        </ul>
        """
    
    return f"""
<html>
<head>
    <style>
        body {{
            font-family: Arial, sans-serif;
            line-height: 1.6;
            color: #333;
            max-width: 600px;
            margin: 0 auto;
            padding: 20px;
        }}
        .greeting {{ margin-bottom: 15px; }}
        .message {{ margin: 15px 0; }}
        .details {{ 
            margin: 20px 0; 
            padding: 15px;
            background-color: #f9f9f9;
            border-left: 4px solid #ff7a59;
        }}
        .details {{{{ margin: 20px 0; padding: 15px; }}}}
        .details p {{{{ margin: 8px 0; }}}}
        .closing {{ margin-top: 15px; }}
        .signature {{ margin-top: 15px; font-weight: bold; }}
        .company {{ color: #666; font-size: 0.9em; }}
    </style>
</head>
<body>
    <div class="greeting">
        <p>Hello {sender_name},</p>
    </div>
    
    <div class="message">
        <p>Here's the {entity_display} that matches your request in HubSpot:</p>
    </div>
    
    <div class="details">
        {details_html}
    </div>
    
    <div class="closing">
        <p>If you need additional information, please don't hesitate to ask.</p>
    </div>
    
    <div class="signature">
        <p>Best regards,<br>
        The HubSpot Assistant Team<br>
        <a href="http://lowtouch.ai" class="company">lowtouch.ai</a></p>
    </div>
</body>
</html>
"""

def build_table_html(sender_name, results, entity, count):
    """Build HTML table for 2-10 results"""
    entity_display = entity.replace("_", " ").title()
    
    # Build table headers and rows based on entity type
    if entity == "contacts":
        headers = ["Contact ID", "Name", "Email", "Phone", "Job Title", "Owner"]
        rows = ""
        for r in results:
            contact_id = r.get('Contact ID', r.get('id', 'N/A'))
            firstname = r.get('Firstname', r.get('firstname', ''))
            lastname = r.get('Lastname', r.get('lastname', ''))
            name = f"{firstname} {lastname}".strip() or "N/A"
            email = r.get('Email', r.get('email', 'N/A'))
            phone = r.get('Phone', r.get('phone', 'N/A'))
            job_title = r.get('Job Title', r.get('jobtitle', 'N/A'))
            owner = r.get('Contact Owner Name', r.get('hubspot_owner_name', 'N/A'))
            
            rows += f"""
            <tr>
                <td>{contact_id}</td>
                <td>{name}</td>
                <td>{email}</td>
                <td>{phone}</td>
                <td>{job_title}</td>
                <td>{owner}</td>
            </tr>
            """
    
    elif entity == "companies":
        headers = ["Company ID", "Company Name", "Domain", "Phone", "Industry"]
        rows = ""
        for r in results:
            company_id = r.get('Company Id', r.get('id', 'N/A'))
            company_name = r.get('Company Name', r.get('name', 'N/A'))
            domain = r.get('Domain', r.get('domain', 'N/A'))
            phone = r.get('Phone', r.get('phone', 'N/A'))
            industry = r.get('Industry', r.get('industry', 'N/A'))
            
            rows += f"""
            <tr>
                <td>{company_id}</td>
                <td>{company_name}</td>
                <td>{domain}</td>
                <td>{phone}</td>
                <td>{industry}</td>
            </tr>
            """
    
    elif entity == "deals":
        headers = ["Deal ID", "Deal Name", "Amount", "Stage", "Close Date", "Owner"]
        rows = ""
        for r in results:
            deal_id = r.get('Deal ID', r.get('id', 'N/A'))
            deal_name = r.get('Deal Name', r.get('dealname', 'N/A'))
            amount = r.get('Deal Amount', r.get('amount', 'N/A'))
            stage = r.get('Deal Stage', r.get('dealstage', 'N/A'))
            close_date = r.get('Expected Close Date', r.get('closedate', 'N/A'))
            owner = r.get('Deal Owner', r.get('hubspot_owner_name', 'N/A'))
            
            # Format amount
            if amount and amount != 'N/A':
                try:
                    amount = f"${int(float(amount)):,}"
                except:
                    pass
            
            rows += f"""
            <tr>
                <td>{deal_id}</td>
                <td>{deal_name}</td>
                <td>{amount}</td>
                <td>{stage}</td>
                <td>{close_date}</td>
                <td>{owner}</td>
            </tr>
            """
    else:
        headers = ["Data"]
        rows = f"<tr><td>{str(results)}</td></tr>"
    
    header_html = "".join([f"<th>{h}</th>" for h in headers])
    
    return f"""
<html>
<head>
    <style>
        body {{
            font-family: Arial, sans-serif;
            line-height: 1.6;
            color: #333;
            max-width: 900px;
            margin: 0 auto;
            padding: 20px;
        }}
        .greeting {{ margin-bottom: 15px; }}
        .message {{ margin: 15px 0; }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
            background: #ffffff;
            border: 1px solid #e0e0e0;
            font-size: 14px;
        }}
        th {{
            background-color: #f3f6fc;
            color: #333;
            padding: 12px;
            border: 1px solid #d0d7e2;
            text-align: left;
            font-weight: bold;
        }}
        td {{
            padding: 10px 12px;
            border: 1px solid #e0e0e0;
            text-align: left;
        }}
        tr:nth-child(even) {{
            background-color: #f9f9f9;
        }}
        .closing {{ margin-top: 15px; }}
        .signature {{ margin-top: 15px; font-weight: bold; }}
        .company {{ color: #666; font-size: 0.9em; }}
    </style>
</head>
<body>
    <div class="greeting">
        <p>Hello {sender_name},</p>
    </div>
    
    <div class="message">
        <p>Here's the {entity_display} from HubSpot based on your request:</p>
    </div>
    
    <table>
        <thead>
            <tr>{header_html}</tr>
        </thead>
        <tbody>
            {rows}
        </tbody>
    </table>
    
    <div class="closing">
        <p>If you need additional information, please don't hesitate to ask.</p>
    </div>
    
    <div class="signature">
        <p>Best regards,<br>
        The HubSpot Assistant Team<br>
        <a href="http://lowtouch.ai" class="company">lowtouch.ai</a></p>
    </div>
</body>
</html>
"""

def build_response_html(sender_name, results, entity, result_count):
    """Build final response HTML with proper field formatting."""
    
    entity_display = entity.replace("_", " ").title()
    
    # Build table based on entity type
    if entity == "contacts":
        headers = ["Contact ID", "Firstname", "Lastname", "Email", "Phone", "Job Title", "Owner", "Last Modified"]
        rows = ""
        for r in results:
            rows += f"""
            <tr>
                <td>{r.get('Contact ID', r.get('id', 'N/A'))}</td>
                <td>{r.get('Firstname', r.get('firstname', 'N/A'))}</td>
                <td>{r.get('Lastname', r.get('lastname', 'N/A'))}</td>
                <td>{r.get('Email', r.get('email', 'N/A'))}</td>
                <td>{r.get('Phone', r.get('phone', 'N/A'))}</td>
                <td>{r.get('Job Title', r.get('jobtitle', 'N/A'))}</td>
                <td>{r.get('Contact Owner Name', r.get('hubspot_owner_id', 'N/A'))}</td>
                <td>{r.get('Last Modified Date', r.get('lastmodifieddate', 'N/A'))}</td>
            </tr>
            """
    elif entity == "companies":
        headers = ["Company ID", "Company Name", "Domain"]
        rows = ""
        for r in results:
            rows += f"""
            <tr>
                <td>{r.get('Company Id', r.get('id', 'N/A'))}</td>
                <td>{r.get('Company Name', r.get('name', 'N/A'))}</td>
                <td>{r.get('Domain', r.get('domain', 'N/A'))}</td>
            </tr>
            """
    elif entity == "deals":
        headers = ["Deal ID", "Deal Name", "Deal Stage", "Owner", "Amount", "Close Date", "Company", "Contacts"]
        rows = ""
        for r in results:
            rows += f"""
            <tr>
                <td>{r.get('Deal ID', r.get('id', 'N/A'))}</td>
                <td>{r.get('Deal Name', r.get('dealname', 'N/A'))}</td>
                <td>{r.get('Deal Stage', r.get('dealstage', 'N/A'))}</td>
                <td>{r.get('Deal Owner', r.get('hubspot_owner_id', 'N/A'))}</td>
                <td>{r.get('Deal Amount', r.get('amount', 'N/A'))}</td>
                <td>{r.get('Expected Close Date', r.get('closedate', 'N/A'))}</td>
                <td>{r.get('Associated Company', r.get('associated_company', 'N/A'))}</td>
                <td>{r.get('Associated Contacts', r.get('associated_contacts', 'N/A'))}</td>
            </tr>
            """
    else:
        headers = ["Data"]
        rows = f"<tr><td>{str(results)}</td></tr>"
    
    header_html = "".join([f"<th>{h}</th>" for h in headers])
    
    html = f"""
<html>
<head>
    <style>
        body {{ font-family: Arial, sans-serif; padding: 20px; max-width: 1200px; margin: 0 auto; }}
        .header {{ background-color: #ff7a59; color: white; padding: 15px; border-radius: 5px; }}
        table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
        th {{ background-color: #f0f0f0; padding: 10px; text-align: left; border: 1px solid #ddd; font-size: 0.85em; }}
        td {{ padding: 10px; border: 1px solid #ddd; font-size: 0.85em; }}
        tr:nth-child(even) {{ background-color: #f9f9f9; }}
        .footer {{ margin-top: 20px; font-size: 0.9em; color: #666; }}
    </style>
</head>
<body>
    <div class="header">
        <h2>✅ {entity_display} Results</h2>
    </div>
    
    <p>Hello {sender_name},</p>
    <p>I found <strong>{result_count} {entity_display}</strong> for you:</p>
    
    <table>
        <tr>{header_html}</tr>
        {rows}
    </table>
    
    <div class="footer">
        <p>Best regards,<br>HubSpot Assistant</p>
    </div>
</body>
</html>
"""
    return html


# ==================== SEND EMAIL WITH ATTACHMENT ====================
def send_email_with_attachment(service, email, html_content, attachment_path, attachment_filename):
    """Send email reply with file attachment."""
    headers = email.get("headers", {})
    
    try:
        # Threading
        original_message_id = headers.get("Message-ID", "")
        references = headers.get("References", "")
        if original_message_id:
            references = f"{references} {original_message_id}".strip() if references else original_message_id

        # Subject
        subject = headers.get("Subject", "No Subject")
        if not subject.lower().startswith("re:"):
            subject = f"Re: {subject}"

        # Recipients
        sender_email = headers.get("From", "")
        
        # Compose
        msg = MIMEMultipart()
        msg["From"] = f"HubSpot via lowtouch.ai <{HUBSPOT_FROM_ADDRESS}>"
        msg["To"] = sender_email
        msg["Subject"] = subject
        if original_message_id:
            msg["In-Reply-To"] = original_message_id
        if references:
            msg["References"] = references
        
        msg.attach(MIMEText(html_content, "html"))
        
        # Attach Excel file
        with open(attachment_path, "rb") as f:
            part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f"attachment; filename={attachment_filename}")
        msg.attach(part)

        # Send
        raw_msg = base64.urlsafe_b64encode(msg.as_string().encode("utf-8")).decode("utf-8")
        service.users().messages().send(userId="me", body={"raw": raw_msg}).execute()
        
        logging.info(f"Sent email with attachment {attachment_filename}")
        return True

    except Exception as e:
        logging.error(f"Failed to send email with attachment: {e}", exc_info=True)
        return False


# ==================== TASK 2: Format Response or Trigger Report ====================
def generate_final_response_or_trigger_report(**kwargs):
    """Final formatting or report trigger. On ANY failure → send technical fallback."""
    ti = kwargs['ti']
    items = ti.xcom_pull(task_ids="analyze_and_search_with_tools", key="llm_search_results") or []
    
    if not items:
        logging.info("No successfully processed emails from Task 1")
        return

    service = authenticate_gmail()
    if not service:
        logging.error("Gmail authentication failed in Task 2")
        return

    for item in items:
        email = item["email"]
        email_id = email.get("id", "unknown")
        sender_name = item["sender_name"]
        analysis = item["analysis"]

        try:
            if not analysis.get("needs_search", False):
                # Casual response
                casual_html = f"""
                <html>
                <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333; max-width: 600px; margin: 0 auto; padding: 20px;">
                    <p>Hello {sender_name},</p>
                    <p>Thank you for your message! We're happy to help with anything HubSpot-related.</p>
                    <p>Feel free to ask about contacts, deals, tasks, or anything else!</p>
                    <p>Best regards,<br>The HubSpot Assistant Team<br>
                    <a href="http://lowtouch.ai" style="color:#666;font-size:0.9em;">lowtouch.ai</a></p>
                </body>
                </html>
                """
                send_email_reply(service, email, casual_html)
                mark_message_as_read(service, email_id)
                continue

            count = analysis.get("result_count", 0)
            logging.info(f"Processing email {email_id} with count: {count}")
            results = analysis.get("results", [])
            logging.info(f"Results for {email_id}: {len(results)} items")
            entity = analysis.get("entity", "records")
            logging.info(f"Entity type for {email_id}: {entity}")

            # Direct routing based on count
            # STRICT COUNT-BASED ROUTING
            logging.info(f"Result count for {email_id}: {count} {entity}")

            if count == 0:
                # No results - use template
                logging.info(f"→ Routing to NO RESULTS template")
                html_content = build_no_results_html(sender_name, entity)
                send_email_reply(service, email, html_content)
                mark_message_as_read(service, email_id)
                continue

            elif count == 1:
                # Single result - inline format
                logging.info(f"→ Routing to SINGLE RESULT (inline) template")
                html_content = build_single_result_html(sender_name, results[0], entity)
                send_email_reply(service, email, html_content)
                mark_message_as_read(service, email_id)
                continue

            elif 2 <= count <= 10:
                # Multiple but manageable - HTML table
                logging.info(f"→ Routing to HTML TABLE template ({count} results)")
                html_content = build_table_html(sender_name, results, entity, count)
                send_email_reply(service, email, html_content)
                mark_message_as_read(service, email_id)
                continue

            else:  # count > 10
                logging.info(f"→ Routing to EXCEL REPORT ({count} results)")
                
                # ✅ PASS THE ALREADY-FOUND RESULTS to report DAG
                email_with_results = email.copy()
                email_with_results["pre_analyzed_results"] = {
                    "results": results,
                    "entity": entity,
                    "count": count,
                    "analysis": analysis  # Pass entire analysis object
                }
    
                ti.xcom_push(key="general_query_report", value=[email_with_results])
                trigger_report_dag(**kwargs)
                continue

            # For count <= 10, generate HTML directly using templates
            prompt = f"""You are a friendly HubSpot email assistant. Generate a professional HTML email response.

Sender: {sender_name}
Entity: {entity}
Results count: {count}

Data:
{json.dumps(results, default=str)}

CRUCIAL VALIDATION INSTRUCTIONS:
**Format response based on count:**
   - 0 results → Use "No Results" template
   - 1 result → Use "Single Result (Inline)" template  
   - 2-10 results → Use "Multiple Results (Table)" template

**HTML TEMPLATE FOR NO RESULTS (0 found):**
<html>
<head>
    <style>
        body {{{{
            font-family: Arial, sans-serif;
            line-height: 1.6;
            color: #333;
            max-width: 600px;
            margin: 0 auto;
            padding: 20px;
        }}}}
        .greeting {{{{ margin-bottom: 15px; }}}}
        .message {{{{ margin: 15px 0; }}}}
        .closing {{{{ margin-top: 15px; }}}}
        .signature {{{{ margin-top: 15px; font-weight: bold; }}}}
        .company {{{{ color: #666; font-size: 0.9em; }}}}
    </style>
</head>
<body>
    <div class="greeting">
        <p>Hello {sender_name},</p>
    </div>
    
    <div class="message">
        <p>I didn't find any {entity} in HubSpot that match the criteria you shared.</p>
    </div>
    
    <div class="closing">
        <p>If you need additional information, please don't hesitate to ask.</p>
    </div>
    
    <div class="signature">
        <p>Best regards,<br>
        The HubSpot Assistant Team<br>
        <a href="http://lowtouch.ai" class="company">lowtouch.ai</a></p>
    </div>
</body>
</html>

**HTML TEMPLATE FOR SINGLE RESULT (1 found - INLINE):**
<html>
<head>
    <style>
        body {{{{
            font-family: Arial, sans-serif;
            line-height: 1.6;
            color: #333;
            max-width: 600px;
            margin: 0 auto;
            padding: 20px;
        }}}}
        .greeting {{{{ margin-bottom: 15px; }}}}
        .message {{{{ margin: 15px 0; }}}}
        .details {{{{ margin: 20px 0; padding: 15px; }}}}
        .details p {{{{ margin: 8px 0; }}}}
        .closing {{{{ margin-top: 15px; }}}}
        .signature {{{{ margin-top: 15px; font-weight: bold; }}}}
        .company {{{{ color: #666; font-size: 0.9em; }}}}
    </style>
</head>
<body>
    <div class="greeting">
        <p>Hello {sender_name},</p>
    </div>
    
    <div class="message">
        <p>Here's the {entity} that matches your request in HubSpot.</p>
    </div>
    
    <div class="details">
        <p><strong>{entity.capitalize()} details:</strong></p>
        <ul>
            <!-- Populate with actual data fields from results[0] -->
        </ul>
    </div>
    
    <div class="closing">
        <p>If you need additional information, please don't hesitate to ask.</p>
    </div>
    
    <div class="signature">
        <p>Best regards,<br>
        The HubSpot Assistant Team<br>
        <a href="http://lowtouch.ai" class="company">lowtouch.ai</a></p>
    </div>
</body>
</html>

**HTML TEMPLATE FOR MULTIPLE RESULTS (2-10 found - TABLE):**
<html>
<head>
    <style>
        body {{{{
            font-family: Arial, sans-serif;
            line-height: 1.6;
            color: #333;
            max-width: 800px;
            margin: 0 auto;
            padding: 20px;
        }}}}
        table {{{{
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
            background: #ffffff;
            border: 1px solid #e0e0e0;
            font-size: 14px;
        }}}}
        th {{{{
            background-color: #f3f6fc;
            color: #333;
            padding: 10px;
            border: 1px solid #d0d7e2;
            text-align: left;
            font-weight: bold;
            white-space: nowrap;
        }}}}
        td {{{{
            padding: 10px;
            border: 1px solid #e0e0e0;
            text-align: left;
            white-space: nowrap;
        }}}}
        .greeting {{{{ margin-bottom: 15px; }}}}
        .message {{{{ margin: 15px 0; }}}}
        .closing {{{{ margin-top: 15px; }}}}
        .signature {{{{ margin-top: 15px; font-weight: bold; }}}}
        .company {{{{ color: #666; font-size: 0.9em; }}}}
    </style>
</head>
<body>
    <div class="greeting">
        <p>Hello {sender_name},</p>
    </div>
    
    <div class="message">
        <p>Here's the list of {entity} from HubSpot based on your request.</p>
    </div>
    
    <table>
        <thead>
            <tr>
                <!-- Populate column headers dynamically based on data fields -->
            </tr>
        </thead>
        <tbody>
            <!-- Populate rows with actual data from results -->
        </tbody>
    </table>
    
    <div class="closing">
        <p>If you need additional information, please don't hesitate to ask.</p>
    </div>
    
    <div class="signature">
        <p>Best regards,<br>
        The HubSpot Assistant Team<br>
        <a href="http://lowtouch.ai" class="company">lowtouch.ai</a></p>
    </div>
</body>
</html>

INSTRUCTIONS:
1. Select the appropriate template based on the count (0, 1, or 2-10)
2. Fill in all placeholders with actual data from the results
3. For single results, use inline formatting with all relevant fields in a bulleted list
4. For multiple results, create a proper table with all relevant columns
5. Keep phone numbers and emails on single lines
6. Use consistent date formatting (MMM DD, YYYY)
7. Return ONLY the complete HTML email - no JSON, no markdown fences, just the HTML
8. Start with <html> and end with </html>"""

            logging.info(f"Requesting HTML generation for {count} {entity}")
            ai_response = get_ai_response(prompt=prompt, expect_json=False)
            logging.info(f"AI response received, length: {len(ai_response)}")

            # Extract HTML from response
            html_content = extract_html_from_response(ai_response)
            
            if not html_content:
                logging.error(f"HTML extraction failed. Raw response preview: {ai_response[:500]}")
                raise ValueError("No HTML extracted from AI response")

            logging.info(f"Successfully extracted HTML ({len(html_content)} chars)")
            send_email_reply(service, email, html_content)
            mark_message_as_read(service, email_id)
            logging.info(f"✓ Response sent successfully for {email_id}")

        except Exception as e:
            logging.warning(f"Task 2 FAILED for {email_id}: {e} → Sending technical fallback")
            raise
        finally:
                mark_message_as_read(service, email_id)


def extract_html_from_response(response):
    """
    Robust HTML extraction from AI response.
    Handles multiple formats: markdown fences, raw HTML, JSON-wrapped HTML.
    """
    if not response:
        return None
    
    response = response.strip()
    
    # Method 1: Extract from ```html markdown fence
    match = re.search(r'```html\s*(.*?)```', response, re.DOTALL | re.IGNORECASE)
    if match:
        html = match.group(1).strip()
        if html.startswith('<html'):
            return html
    
    # Method 2: Extract from generic ``` fence (in case no language specified)
    match = re.search(r'```\s*(<!DOCTYPE html>.*?</html>|<html.*?</html>)', response, re.DOTALL | re.IGNORECASE)
    if match:
        return match.group(1).strip()
    
    # Method 3: Find complete <html>...</html> block
    match = re.search(r'(<html[^>]*>.*?</html>)', response, re.DOTALL | re.IGNORECASE)
    if match:
        return match.group(1).strip()
    
    # Method 4: Find <!DOCTYPE html>...</html> block
    match = re.search(r'(<!DOCTYPE html>.*?</html>)', response, re.DOTALL | re.IGNORECASE)
    if match:
        return match.group(1).strip()
    
    # Method 5: Check if entire response is HTML (starts with <html or <!DOCTYPE)
    if response.startswith('<!DOCTYPE') or response.startswith('<html'):
        return response
    
    # Method 6: Try to extract from JSON if response is JSON-wrapped
    try:
        json_obj = json.loads(response)
        if isinstance(json_obj, dict):
            # Check common keys that might contain HTML
            for key in ['html', 'html_content', 'content', 'response', 'email_html']:
                if key in json_obj and isinstance(json_obj[key], str):
                    html = json_obj[key].strip()
                    if '<html' in html.lower():
                        return html
    except:
        pass
    
    logging.warning("HTML extraction failed with all methods")
    return None

def extract_and_parse_json(text: str):
    """
    Robustly extract and parse JSON from AI response text.
    Handles: markdown fences, extra text, thinking tags, multiple JSON objects, etc.
    Returns parsed JSON or raises informative error.
    """
    if not text or not text.strip():
        raise ValueError("Empty response from AI")

    # Step 1: Remove thinking tags
    text = re.sub(r'<think>.*?</think>', '', text, flags=re.DOTALL | re.IGNORECASE)

    # Step 2: Remove common markdown fences
    text = re.sub(r'```json\s*', '', text, flags=re.IGNORECASE)
    text = re.sub(r'```\s*', '', text, flags=re.IGNORECASE)

    # Step 3: Strip any leading/trailing whitespace and text
    text = text.strip()

    # Step 4: Find the first valid JSON object or array
    # Look for { ... } or [ ... ]
    json_match = re.search(r'(\{.*\}|\[.*\])', text, re.DOTALL)
    
    if not json_match:
        # Log some context for debugging
        preview = text[:200].replace('\n', ' ')
        raise ValueError(f"No JSON object found in AI response. Preview: '{preview}'")

    json_str = json_match.group(1)

    # Step 5: Basic cleanup of common issues inside JSON
    # Remove trailing commas before } or ] (common LLM mistake)
    json_str = re.sub(r',\s*}', '}', json_str)
    json_str = re.sub(r',\s*]', ']', json_str)

    try:
        return json.loads(json_str)
    except json.JSONDecodeError as e:
        # Final attempt: try to fix unbalanced braces (rare but happens)
        raise ValueError(f"Invalid JSON after extraction: {e}\nExtracted string preview: {json_str[:300]}")

def get_deal_stage_labels():
    """
    Fetch all pipelines and stages from HubSpot and return a mapping:
    { 'appointmentscheduled': 'Appointment Scheduled', ... }
    """
    try:
        endpoint = f"{HUBSPOT_BASE_URL}/crm/v3/pipelines/deals"
        headers = {
            "Authorization": f"Bearer {HUBSPOT_API_KEY}",
            "Content-Type": "application/json"
        }
        
        response = requests.get(endpoint, headers=headers, timeout=30)
        response.raise_for_status()
        pipelines_data = response.json().get("results", [])

        stage_mapping = {}
        for pipeline in pipelines_data:
            pipeline_label = pipeline.get("label", "")
            stages = pipeline.get("stages", [])
            for stage in stages:
                stage_id = stage.get("id")
                stage_label = stage.get("label")
                if stage_id and stage_label:
                    stage_mapping[stage_id] = stage_label
                    # Optional: include pipeline name for context
                    # stage_mapping[stage_id] = f"{stage_label} ({pipeline_label})"

        logging.info(f"Loaded {len(stage_mapping)} deal stage labels from HubSpot")
        return stage_mapping

    except Exception as e:
        logging.error(f"Failed to fetch deal stage labels: {e}")
        return {}

def build_report_log(entity_type, filters, results_data, requester_email, 
                     original_query, timezone_str="Asia/Kolkata"):
    """
    Build report log JSON for logging purposes only (no disk write).
    
    Args:
        entity_type (str): Type of entity (deals, contacts, companies)
        filters (list): Applied filters from search
        results_data (list): Query results from HubSpot
        requester_email (str): Email of person requesting report
        original_query (str): Original user query text
        timezone_str (str): Timezone for timestamps
        
    Returns:
        dict: Report log structure for logging
    """
    try:
        # Generate unique report ID
        report_id = str(uuid.uuid4())
        
        # Get timezone-aware timestamps
        tz = pytz.timezone(timezone_str)
        request_time = datetime.now(tz)
        data_as_of = datetime.now(tz)
        
        # Format timestamps as ISO 8601 with timezone
        request_timestamp = request_time.isoformat()
        data_as_of_timestamp = data_as_of.isoformat()
        
        # Extract requester_id from email
        requester_id = f"user:{requester_email}"
        
        # Build filter representation
        filter_dict = {}
        for f in filters:
            prop_name = f.get("propertyName")
            operator = f.get("operator")
            value = f.get("value")
            
            # Convert to report-friendly format
            if operator == "GT":
                filter_dict[f"{prop_name}_gt"] = value
            elif operator == "GTE":
                filter_dict[f"{prop_name}_from"] = value
            elif operator == "LT":
                filter_dict[f"{prop_name}_lt"] = value
            elif operator == "LTE":
                filter_dict[f"{prop_name}_to"] = value
            elif operator == "EQ":
                filter_dict[prop_name] = value
            else:
                filter_dict[f"{prop_name}_{operator.lower()}"] = value
        
        # Format results based on entity type
        formatted_results = []
        total_value = 0
        highest_deal = None
        
        for record in results_data:
            props = record.get("properties", {})
            record_id = record.get("id")
            
            if entity_type == "deals":
                amount = float(props.get("amount", 0) or 0)
                formatted_record = {
                    "dealId": record_id,
                    "dealName": props.get("dealname", ""),
                    "amount": amount,
                    "close_date": props.get("closedate", ""),
                    "stage": props.get("dealstage", ""),
                    "pipeline": props.get("pipeline", ""),
                    "ownerName": props.get("hubspot_owner_name", "")
                }
                
                total_value += amount
                if highest_deal is None or amount > highest_deal["amount_converted"]:
                    highest_deal = {
                        "dealId": record_id,
                        "amount_converted": amount
                    }
                    
            elif entity_type == "contacts":
                formatted_record = {
                    "contactId": record_id,
                    "firstName": props.get("firstname", ""),
                    "lastName": props.get("lastname", ""),
                    "email": props.get("email", ""),
                    "phone": props.get("phone", ""),
                    "jobTitle": props.get("jobtitle", "")
                }
                
            elif entity_type == "companies":
                formatted_record = {
                    "companyId": record_id,
                    "companyName": props.get("name", ""),
                    "domain": props.get("domain", ""),
                    "industry": props.get("industry", ""),
                    "phone": props.get("phone", "")
                }
            
            formatted_results.append(formatted_record)
        
        # Build aggregates
        aggregates = {}
        if entity_type == "deals":
            aggregates = {
                "total_deals": len(formatted_results),
                "total_value_converted": total_value,
                "highest_deal": highest_deal
            }
        else:
            aggregates = {
                f"total_{entity_type}": len(formatted_results)
            }
        
        # Build complete report log
        report_log = {
            "report_id": report_id,
            "report_type": f"hubspot_{entity_type}",
            "requester_id": requester_id,
            "request_timestamp": request_timestamp,
            "data_as_of": data_as_of_timestamp,
            "filters": filter_dict,
            "results_count": len(formatted_results),
            "results": formatted_results,
            "aggregates": aggregates,
            "pagination": {
                "page_size": 500,
                "next_cursor": None
            },
            "partial": False,
            "failures": []
        }
        
        return report_log
        
    except Exception as e:
        logging.error(f"Failed to build report log: {e}", exc_info=True)
        return None

def trigger_report_dag(**kwargs):
    """Enhanced trigger_report_dag with professional report email formatting"""
    DEAL_STAGE_LABELS = get_deal_stage_labels()
    ti = kwargs['ti']
    report_emails = ti.xcom_pull(key="report_emails") or []
    general_report_emails = ti.xcom_pull(key="general_query_report") or []
    
    all_report_emails = report_emails + general_report_emails
    
    if not all_report_emails:
        logging.info("No report emails to process (both XCom keys empty)")
        return
    
    logging.info(f"Processing {len(all_report_emails)} report email(s) from general query")
   
    service = authenticate_gmail()
    if not service:
        logging.error("Gmail authentication failed, cannot send reports")
        return
   
    owner_cache = {}
   
    def get_owner_name(owner_id):
        if not owner_id:
            return "Unassigned"
        if owner_id in owner_cache:
            return owner_cache[owner_id]
       
        try:
            url = f"{HUBSPOT_BASE_URL}/crm/v3/owners/{owner_id}"
            headers = {"Authorization": f"Bearer {HUBSPOT_API_KEY}"}
            resp = requests.get(url, headers=headers, timeout=10)
            if resp.status_code == 200:
                data = resp.json()
                name = f"{data.get('firstName', '')} {data.get('lastName', '')}".strip()
                owner_cache[owner_id] = name or "Unknown Owner"
                return owner_cache[owner_id]
            else:
                owner_cache[owner_id] = "Unknown Owner"
                return "Unknown Owner"
        except Exception as e:
            logging.warning(f"Failed to fetch owner {owner_id}: {e}")
            owner_cache[owner_id] = "Unknown Owner"
            return "Unknown Owner"
   
    def format_currency(amount):
        """Format amount as currency"""
        if not amount:
            return "$0"
        try:
            return f"${int(float(amount)):,}"
        except:
            return f"${amount}"
   
    def get_filter_summary(filters, entity_type):
        """Generate human-readable filter summary"""
        if not filters:
            return "No filters applied"
       
        summaries = []
        for f in filters:
            prop = f.get("propertyName", "")
            op = f.get("operator", "")
            val = f.get("value", "")
           
            # Readable operator names
            op_map = {
                "EQ": "equals",
                "NEQ": "not equals",
                "GT": "greater than",
                "GTE": "greater than or equal to",
                "LT": "less than",
                "LTE": "less than or equal to",
                "CONTAINS_TOKEN": "contains",
                "NOT_CONTAINS_TOKEN": "does not contain"
            }
           
            # Special handling for deal stages
            if prop == "dealstage" and entity_type == "deals":
                val = DEAL_STAGE_LABELS.get(val, val)
           
            summaries.append(f"{prop} {op_map.get(op, op)} {val}")
       
        return " • ".join(summaries)
   
    for email in all_report_emails:
        try:
            headers = email.get("headers", {})
            sender_email = headers.get("From", "")
            email_id = email.get("id", "unknown")
            email_match = re.search(r'<(.+?)>', sender_email)
            clean_sender_email = email_match.group(1) if email_match else sender_email.lower()
            chat_history = email.get("chat_history", [])

            # === SPELLING VARIANT DETECTION FOR REPORTS ===
            user_message = email.get("content", "").strip()
            spelling_inject_text, spelling_variants = generate_spelling_variants_for_prompt(
                user_content=user_message,
                chat_history=chat_history
            )
            sender_name = "there"
            name_match = re.search(r'^([^<]+)', sender_email)
            if name_match:
                sender_name = name_match.group(1).strip()
           
            ai_response = None
            report_filepath = None
            report_filename = None
            report_success = False
           
            try:
                # Get conversation history
                conversation_history_for_ai = []
                chat_history = email.get("chat_history", [])
                for i in range(0, len(chat_history), 2):
                    if i + 1 < len(chat_history):
                        user_msg = chat_history[i]
                        assistant_msg = chat_history[i + 1]
                        if user_msg["role"] == "user" and assistant_msg["role"] == "assistant":
                            conversation_history_for_ai.append({
                                "prompt": user_msg["content"],
                                "response": assistant_msg["content"]
                            })
               
                # AI Analysis
                analysis_prompt = f"""You are a HubSpot data analyst. Analyze this request and determine what data to search for.
User request: "{email.get("content", "").strip()}"
extract the enitites details also use the spelling variants:{spelling_inject_text}
- Always generate a report whenever the user explicitly requests it using phrases like "create a report", "generate a report", "give the report", "provide the report", "prepare a report", "build a report", or "produce a report", regardless of record count or thresholds.
IMPORTANT RULES:
1. For company-related queries (e.g., "deals associated with company XYZ"):
   - Use the "associations.company" property to filter by company id
   - Parse the company name and search the company using search_company tool and get the id for the filters.
   - Operator should be CONTAINS_TOKEN for partial matches
   - Example: {{"propertyName": "associations.company", "operator": "CONTAINS_TOKEN", "value": "123"}}

2. For contact-related queries (e.g., "deals for contact John Doe"):
   - Use the "associations.contact" property to filter by contact id
   -  Parse the contact name and search the contact using search_contact tool and get the id for the filters. ALways use
   - Example: {{"propertyName": "associations.contact", "operator": "CONTAINS_TOKEN", "value": "123"}}

3. For date ranges:
   - Always use YYYY-MM-DD format
   - Use GTE for "from" dates and LTE for "to" dates
   - For "this month", use GTE=today and LTE=end_of_month
   - NEVER include past dates before today unless explicitly requested

4. For deal properties:
   - dealstage: Use exact stage IDs (e.g., "appointmentscheduled")
   - dealname: Use CONTAINS_TOKEN for partial matches
   - amount: Use GT/GTE/LT/LTE for range queries
   - closedate: Use GT/GTE/LT/LTE for date ranges
   - open deals include the deals all those deals whose deal stages are not closed won or closed lost.

5. For contact properties:
   - Include hubspot_owner_id in properties list
   - email, firstname, lastname: Use CONTAINS_TOKEN for partial matches

6. Report title should describe what data is shown (e.g., "Deals Associated with Company XYZ")

7. CRITICAL - Deal Stage Mapping:
   - When user mentions a deal stage name (like "lead", "qualified", "closed won"), you MUST:
     a) First call the get_deal_stage_labels() function to get the stage ID mapping
     b) Find the matching stage ID (e.g., "lead" might be "appointmentscheduled")
     c) Use the STAGE ID in the filter, not the human-readable name
   - Example: If user says "lead stage", search for the ID that maps to "Lead"
   - The stage IDs are already loaded in DEAL_STAGE_LABELS variable

Return ONLY a valid JSON object:
{{
    "entity_type": "deals|contacts|companies",
    "filters": [
        {{"propertyName": "property_name", "operator": "EQ|NEQ|GT|GTE|LT|LTE|CONTAINS_TOKEN|NOT_CONTAINS_TOKEN", "value": "filter_value"}}
    ],
    "properties": ["list", "of", "properties"],
    "sort": {{"propertyName": "property_name", "direction": "ASCENDING|DESCENDING"}},
    "report_title": "Descriptive Report Title"
}}
Supported operators: EQ, NEQ, LT, LTE, GT, GTE, CONTAINS_TOKEN, NOT_CONTAINS_TOKEN
"""
               
                analysis_response = get_ai_response(
                    prompt=analysis_prompt,
                    conversation_history=conversation_history_for_ai,
                    expect_json=True
                )
                logging.info(f"AI response is:{analysis_response}")
               
                try:
                    analysis_data = json.loads(analysis_response)
                except:
                    analysis_data = extract_json_from_text(analysis_response)
               
                if not analysis_data or "entity_type" not in analysis_data:
                    raise ValueError("Invalid analysis response from AI")
               
                entity_type = analysis_data.get("entity_type", "contacts").lower()
                filters = analysis_data.get("filters", [])
                properties = analysis_data.get("properties", [])
                sort = analysis_data.get("sort")
                report_title = analysis_data.get("report_title", "HubSpot Report")

                if entity_type == "deals" and filters:
                    DEAL_STAGE_LABELS = get_deal_stage_labels()
                    LABEL_TO_ID = {v.lower(): k for k, v in DEAL_STAGE_LABELS.items()}
                    
                    logging.info(f"Available deal stages: {DEAL_STAGE_LABELS}")
                    logging.info(f"Label to ID mapping: {LABEL_TO_ID}")
                    
                    for filter_obj in filters:
                        if filter_obj.get("propertyName") == "dealstage":
                            stage_value = filter_obj.get("value", "").lower()
                            logging.info(f"Original stage value from AI: '{stage_value}'")
                            
                            # Check if it's already a valid stage ID
                            if stage_value in DEAL_STAGE_LABELS:
                                logging.info(f"✓ '{stage_value}' is already a valid stage ID")
                            # Try to convert label to ID
                            elif stage_value in LABEL_TO_ID:
                                matching_id = LABEL_TO_ID[stage_value]
                                logging.info(f"✓ Converted stage label '{stage_value}' to ID '{matching_id}'")
                                filter_obj["value"] = matching_id
                            else:
                                # Try partial matching
                                found = False
                                for label_lower, stage_id in LABEL_TO_ID.items():
                                    if stage_value in label_lower or label_lower in stage_value:
                                        logging.info(f"✓ Partial match: '{stage_value}' matched to '{label_lower}' (ID: {stage_id})")
                                        filter_obj["value"] = stage_id
                                        found = True
                                        break
                                
                                if not found:
                                    logging.warning(f"⚠️ Could not find stage ID for '{stage_value}'")
                                    logging.warning(f"Available stage labels: {list(LABEL_TO_ID.keys())}")
                                    logging.warning(f"Available stage IDs: {list(DEAL_STAGE_LABELS.keys())}")

                logging.info(f"Final filters being sent to HubSpot: {filters}")

            # ✅ FIX OWNER FILTERS - Convert owner names to IDs using ALL spelling variants
                for filter_obj in filters:
                    if filter_obj.get("propertyName") == "hubspot_owner_id":
                        owner_value = filter_obj.get("value", "")
                        
                        # Check if value looks like a name (not a numeric ID)
                        if not owner_value.isdigit():
                            logging.info(f"Owner filter contains name '{owner_value}', resolving to ID...")
                            
                            # ✅ BUILD LIST OF ALL VARIANTS TO SEARCH
                            search_variants = [owner_value.lower().strip()]  # Start with original
                            
                            # Add spelling variants if available
                            if spelling_variants:
                                for entity_type in ['contacts', 'companies', 'deals']:
                                    variants_list = spelling_variants.get(entity_type, [])
                                    for variant_group in variants_list:
                                        original = variant_group.get('original', '').lower()
                                        variants = variant_group.get('variants', [])
                                        
                                        # If this variant group matches our owner name, add all variants
                                        if original == owner_value.lower():
                                            search_variants.extend([v.lower() for v in variants if v.lower() not in search_variants])
                                            logging.info(f"✅ Added spelling variants for owner '{owner_value}': {variants}")
                                            break
                            
                            logging.info(f"🔍 Searching for owner using variants: {search_variants}")
                            
                            try:
                                # Get all owners from HubSpot
                                endpoint = f"{HUBSPOT_BASE_URL}/crm/v3/owners"
                                headers_req = {
                                    "Authorization": f"Bearer {HUBSPOT_API_KEY}",
                                    "Content-Type": "application/json"
                                }
                                
                                resp = requests.get(endpoint, headers=headers_req, timeout=30)
                                resp.raise_for_status()
                                owners = resp.json().get("results", [])
                                
                                # ✅ SEARCH ALL VARIANTS UNTIL WE FIND A MATCH
                                found_owner_id = None
                                found_owner_name = None
                                matched_variant = None
                                
                                for search_term in search_variants:
                                    if found_owner_id:
                                        break  # Already found a match
                                    
                                    for owner in owners:
                                        first_name = owner.get("firstName", "")
                                        last_name = owner.get("lastName", "")
                                        full_name = f"{first_name} {last_name}".strip()
                                        owner_email = owner.get("email", "")
                                        current_owner_id = owner.get("id")
                                        
                                        # Check if search term matches (case-insensitive)
                                        if (search_term in full_name.lower() or 
                                            search_term in first_name.lower() or 
                                            search_term in last_name.lower() or 
                                            search_term in owner_email.lower() or
                                            full_name.lower() == search_term or
                                            first_name.lower() == search_term or
                                            last_name.lower() == search_term):
                                            
                                            found_owner_id = current_owner_id
                                            found_owner_name = full_name
                                            matched_variant = search_term
                                            logging.info(f"✅ MATCH FOUND using variant '{search_term}'!")
                                            logging.info(f"   Owner: {full_name} ({owner_email}) → ID: {current_owner_id}")
                                            break
                                
                                if found_owner_id:
                                    # Update filter with owner ID
                                    filter_obj["value"] = found_owner_id
                                    filter_obj["operator"] = "EQ"  # Change to exact match
                                    logging.info(f"✅ Owner resolved: '{owner_value}' → '{found_owner_name}' (ID: {found_owner_id})")
                                    logging.info(f"   Matched using variant: '{matched_variant}'")
                                else:
                                    logging.warning(f"⚠️ No owner found matching '{owner_value}' or any of its variants: {search_variants}")
                                    
                            except Exception as e:
                                logging.error(f"Failed to resolve owner name '{owner_value}': {e}", exc_info=True)
                        else:
                            logging.info(f"✓ Owner filter already contains ID: {owner_value}")

                        logging.info(f"📤 Final filters after owner resolution: {filters}")

                # Default properties
                base_properties = {
                    "contacts": ['firstname', 'lastname', 'email', 'phone', 'jobtitle', 'address', 'city', 'state', 'country', 'hs_object_id', 'hubspot_owner_id'],
                    "companies": ['name', 'domain', 'phone', 'address', 'city', 'state', 'zip', 'country', 'industry', 'type', 'hs_object_id'],
                    "deals": ['dealname', 'amount', 'closedate', 'dealstage', 'pipeline', 'hubspot_owner_id', 'hs_object_id']
                }
               
                default_props = base_properties.get(entity_type, [])
                final_properties = list(set(default_props + (properties or [])))
               
                # Execute search
                # Execute search - FORCE FETCH ALL RESULTS (no max_results limit)
                logging.info(f"Executing search for {entity_type} with filters: {filters}")

                if entity_type == "contacts":
                    search_results = search_contacts(
                        filters=filters, 
                        properties=final_properties,
                        limit=200,  # Max per page
                        max_results=None  # No limit - fetch everything
                    )
                elif entity_type == "companies":
                    search_results = search_companies(
                        filters=filters, 
                        properties=final_properties,
                        limit=200,
                        max_results=None
                    )
                elif entity_type == "deals":
                    search_results = search_deals(
                        filters=filters, 
                        properties=final_properties,
                        limit=200,
                        max_results=None
                    )

                logging.info(f"Search completed: {len(search_results.get('results', []))} total {entity_type} found")
               
                results_data = search_results.get("results", [])
               
                if not results_data:
                    error_msg = search_results.get("error", "")
                    logging.info(f"⚠️ No results found for {entity_type}. HubSpot message: {error_msg}")
                    
                    # Set flag and prepare user-friendly response
                    zero_results = True
                    report_success = True  # Not a technical failure, just no data
                    
                    # Build filter summary for the message
                    filter_summary = get_filter_summary(filters, entity_type)
                    
                    # Get timezone for timestamp
                    timezone_str = "Asia/Kolkata"
                    tz = pytz.timezone(timezone_str)
                    current_time = datetime.now(tz)
                    data_as_of_formatted = current_time.strftime("%b %d, %Y")
                    
                    # Create friendly "no results" response
                    ai_response = f"""
<html>
<head>
    <style>
        body {{
            font-family: Arial, 'Helvetica Neue', Helvetica, sans-serif;
            line-height: 1.6;
            color: #000000;
            max-width: 600px;
            margin: 0 auto;
            padding: 20px;
        }}
        .greeting {{
            margin-bottom: 8px;
        }}
        .report-title {{
            font-size: 15px;
            font-weight: bold;
            color: #000000;
            margin: 8px 0 8px 0;
            padding-bottom: 8px;
        }}
        .metadata {{
            font-size: 13px;
            color: #333333;
            margin: 8px 0 20px 0;
        }}
        .no-results-title {{
            font-weight: bold;
            color: #000000;
            margin-bottom: 6px;
        }}
        .filter-info {{
            margin: 12px 0;
            font-size: 13px;
            color: #000000;
        }}
        .suggestions {{
            margin: 20px 0;
        }}
        .suggestions ul {{
            margin: 10px 0;
            padding-left: 20px;
        }}
        .suggestions li {{
            margin: 8px 0;
        }}
        .message {{
            margin: 8px 0;
            color: #000000;
        }}
        .closing {{
            margin-top: 8px;
            color: #000000;
        }}
        .signature {{
            margin-top: 30px;
            padding-top: 15px;
        }}
        .signature-line {{
            margin: 3px 0;
            color: #000000;
        }}
        .company {{
            color: #666666;
            font-size: 13px;
            margin-top: 8px;
        }}
        .company a {{
            color: #0066cc;
            text-decoration: none;
        }}
    </style>
</head>
<body>

    <div class="greeting">
        <p>Hello {sender_name},</p>
    </div>

    <div class="report-title">
        {report_title}
    </div>

    <div class="metadata">
        Data as of: {data_as_of_formatted}
    </div>

    <div>
        <div class="no-results-title">No Matching Results Found</div>
        <p>We couldn’t find any {entity_type} that match the specified criteria.</p>
    </div>

    <div class="filter-info">
        <strong>Applied Filters:</strong><br>
        {filter_summary}
    </div>

    <div class="suggestions">
        <strong>Suggestions:</strong>
        <ul>
            <li>Consider broadening your search criteria or date range.</li>
            <li>Verify that the applied filters are accurate.</li>
            <li>Check whether the relevant deal data exists in HubSpot.</li>
            <li>Contact support if you believe the expected data should appear.</li>
        </ul>
    </div>

    <div class="message">
        <p>If you’d like to adjust your search criteria or need help retrieving specific information, feel free to reply to this email with your updated requirements.</p>
    </div>

    <div class="closing">
        <p>I’m here to assist you further.</p>
    </div>

    <div class="signature">
        <div class="signature-line"><strong>Best regards,</strong></div>
        <div class="signature-line">The HubSpot Assistant Team</div>
        <div class="company"><a href="http://lowtouch.ai">lowtouch.ai</a></div>
    </div>

</body>
</html>
"""
                    
                    # Log for tracking (but not as a failure)
                    zero_results_log = {
                        "report_id": str(uuid.uuid4()),
                        "report_type": f"hubspot_{entity_type}",
                        "requester_id": f"user:{clean_sender_email}",
                        "request_timestamp": current_time.isoformat(),
                        "results_count": 0,
                        "filters": filter_summary,
                        "status": "completed_no_results"
                    }
                    logging.info(f"REPORT_NO_RESULTS: {json.dumps(zero_results_log)}")
                else:
                    logging.info(f"✅ Retrieved {len(results_data)} {entity_type} from HubSpot")
                    report_log = build_report_log(
                        entity_type=entity_type,
                        filters=filters,
                        results_data=results_data,
                        requester_email=clean_sender_email,
                        original_query=email.get("content", "").strip(),
                        timezone_str="Asia/Kolkata"
                    )

                    if report_log:
                        # Log the report metadata as structured JSON
                        logging.info(f"REPORT_GENERATED: {json.dumps(report_log, indent=2)}")
                        
                        # Extract report_id for reference
                        report_id = report_log.get("report_id")
                        logging.info(f"✓ Report generated: report_id={report_id}, "
                                f"entity_type={entity_type}, results_count={len(results_data)}")
                    else:
                        logging.warning("Failed to build report log")

                
                    # Calculate summary statistics
                    record_count = len(results_data)
                    total_value = 0
                    unique_owners = set()
                
                    # Format data with helper columns
                    formatted_data = []
                    for record in results_data:
                        props = record.get("properties", {})
                        record_id = record.get("id")
                    
                        row = {"_record_id": record_id}
                    
                        if entity_type == "contacts":
                            full_name = f"{props.get('firstname', '')} {props.get('lastname', '')}".strip()
                            owner_name = get_owner_name(props.get("hubspot_owner_id"))
                            if props.get("hubspot_owner_id"):
                                unique_owners.add(props.get("hubspot_owner_id"))
                        
                            row.update({
                                "Contact ID": record_id,
                                "Contact Name": full_name or "No Name",
                                "Job Title": props.get("jobtitle", ""),
                                "Email": props.get("email", ""),
                                "Phone": props.get("phone", ""),
                                "Street Address": props.get("address", ""),
                                "City": props.get("city", ""),
                                "State/Region": props.get("state", ""),
                                "Postal Code": props.get("zip", ""),
                                "Country": props.get("country", ""),
                                "Contact Owner": owner_name,
                                "_hubspot_url": f"{HUBSPOT_UI_URL}/contact/{record_id}"
                            })
                    
                        elif entity_type == "companies":
                            row.update({
                                "Company ID": record_id,
                                "Company Name": props.get("name", "Unnamed Company"),
                                "Domain": props.get("domain", ""),
                                "Street Address": props.get("address", ""),
                                "City": props.get("city", ""),
                                "State/Region": props.get("state", ""),
                                "Country": props.get("country", ""),
                                "Phone": props.get("phone", ""),
                                "Type": props.get("type", ""),
                                "_hubspot_url": f"{HUBSPOT_UI_URL}/company/{record_id}"
                            })
                    
                        elif entity_type == "deals":
                            owner_name = get_owner_name(props.get("hubspot_owner_id"))
                            if props.get("hubspot_owner_id"):
                                unique_owners.add(props.get("hubspot_owner_id"))
                        
                            stage_id = props.get("dealstage")
                            stage_label = DEAL_STAGE_LABELS.get(stage_id, "Unknown Stage")
                        
                            # Add to total value
                            amount = props.get("amount")
                            if amount:
                                try:
                                    total_value += float(amount)
                                except:
                                    pass

                            close_date_value = props.get("closedate")
                            if close_date_value:
                                try:
                                    # Check if it's a timestamp string like "2025-12-02T00:00:00Z"
                                    if isinstance(close_date_value, str) and ('T' in close_date_value or '-' in close_date_value):
                                        close_date_dt = parser.parse(close_date_value).replace(tzinfo=None)
                                    else:
                                        # It's a millisecond timestamp
                                        close_date_dt = datetime.fromtimestamp(int(close_date_value) / 1000)
                                except:
                                    close_date_dt = None
                            else:
                                close_date_dt = None
                        
                            row.update({
                                "Deal ID": record_id,
                                "Deal Name": props.get("dealname", "Untitled Deal"),
                                "Amount": props.get("amount", ""),
                                "Close Date": close_date_dt,
                                "Deal Stage": stage_label,
                                "Deal Owner": owner_name,
                                "_hubspot_url": f"{HUBSPOT_UI_URL}/deal/{record_id}"
                            })
                    
                        formatted_data.append(row)
                
                    # Column order
                    column_order = {
                        "contacts": ["Contact ID", "Contact Name", "Job Title", "Email", "Phone", "Street Address", "City", "State/Region", "Country", "Contact Owner"],
                        "companies": ["Company ID", "Company Name", "Domain", "Street Address", "City", "State/Region", "Country", "Phone", "Type"],
                        "deals": ["Deal ID", "Deal Name", "Amount", "Deal Stage", "Close Date", "Deal Owner"]
                    }
                
                    df = pd.DataFrame(formatted_data)
                    ordered_cols = column_order.get(entity_type, df.columns.tolist())
                
                    # Reorder only visible columns, keep helper columns
                    visible_cols = [col for col in ordered_cols if col in df.columns]
                    df_final = df[visible_cols + ['_hubspot_url']]
                
                    # Make sure the ID column is first
                    id_col = "Deal ID" if entity_type == "deals" else \
                            "Contact ID" if entity_type == "contacts" else "Company ID"
                
                    if id_col in df_final.columns:
                        cols = [id_col] + [c for c in df_final.columns if c != id_col and c != '_hubspot_url'] + ['_hubspot_url']
                        df_final = df_final[cols]
                
                    # Export with hyperlinks
                    export_result = export_to_file(
                        data=df_final,
                        export_format='excel',
                        filename=f"hubspot_{entity_type}_report",
                        export_dir='/appz/cache/exports',
                        hyperlinks={
                            'Sheet1': {
                                id_col: {
                                    'url_column': '_hubspot_url'
                                }
                            }
                        }
                    )
                
                    if not export_result.get("success"):
                        raise ValueError(f"Export failed: {export_result.get('error', 'Unknown error')}")
                
                    report_filepath = export_result["filepath"]
                    report_filename = export_result["filename"]
                    report_success = True
                
                    logging.info(f"✓ Report generated successfully: {report_filepath}")
                
                    timezone_str = "Asia/Kolkata"
                    tz = pytz.timezone(timezone_str)
                    current_time = datetime.now(tz)
                    # Format date as dd-Mon-yyyy (e.g., 03-Dec-2025)
                    data_as_of_formatted = current_time.strftime("%b %d, %Y")
                
                    # Build filter summary
                    filter_summary = get_filter_summary(filters, entity_type)
                
                    # Owner summary
                    owner_summary = f"{len(unique_owners)} owner(s)" if unique_owners else "All owners"
                
                    # Success HTML - Professional, clean format
                    ai_response = f"""
    <html>
    <head>
        <style>
            body {{
                font-family: Arial, 'Helvetica Neue', Helvetica, sans-serif;
                line-height: 1.6;
                color: #000000;
                max-width: 600px;
                margin: 0 auto;
                padding: 20px;
            }}
            .greeting {{
                margin-bottom: 8px;
            }}
            .report-title {{
                font-size: 15px;
                font-weight: bold;
                color: #000000;
                margin: 8px 0 10px 0;
                padding-bottom: 8px;
            }}
            .metadata {{
                font-size: 13px;
                color: #333333;
                margin: 8px 0 20px 0;
            }}
            .stats-section {{
                margin: 8px;
            }}
            .stat-row {{
                display: flex;
                justify-content: space-between;
                padding: 0;       /* remove extra height */
                margin: 13px 0;
            }}
            .stat-label {{
                font-weight: 600;
                color: #000000;
            }}
            .stat-value {{
                color: #000000;
                font-weight: bold;
                margin-left: 8px
            }}
            .message {{
                margin: 8px 0;
                color: #000000;
            }}
            .closing {{
                margin-top: 8px;
                color: #000000;
            }}
            .signature {{
                margin-top: 30px;
                padding-top: 15px;
            }}
            .signature-line {{
                margin: 3px 0;
                color: #000000;
            }}
            .company {{
                color: #666666;
                font-size: 13px;
                margin-top: 8px;
            }}
            .company a {{
                color: #0066cc;
                text-decoration: none;
            }}
        </style>
    </head>
    <body>
        <div class="greeting">
            <p>Hello {sender_name},</p>
        </div>
    
        <div class="report-title">
            {report_title}
        </div>
        <div class="metadata">
            Data as of: {data_as_of_formatted}
        </div>
    
        <div class="summary" style="margin: 30px 0;">
            <h3 style="margin: 0 0 15px 0; color: #000000;">Summary</h3>
            <ul style="margin: 10px 0; padding-left: 20px;">
                <li><strong>Total {entity_type}:</strong> {record_count}</li>
                {""
                if entity_type.lower() != "deals" and entity_type.lower() != "deal" else
                f'<li><strong>Total Value:</strong>{format_currency(total_value)}</li>'
                }
            </ul>
        </div>
    
        <div class="message">
            <p>Please find the attached Excel file containing your detailed report.</p>
        </div>
    
        <div class="closing">
            <p>If you need any further assistance or have questions regarding the results, please feel free to let me know.</p>
        </div>
    
        <div class="signature">
            <div class="signature-line"><strong>Best regards,</strong></div>
            <div class="signature-line">The HubSpot Assistant Team</div>
            <div class="company"><a href="http://lowtouch.ai">lowtouch.ai</a>
            </div>
        </div>
    </body>
    </html>
    """
           
            except Exception as report_error:
                logging.error(f"Report generation failed for email {email_id}: {report_error}", exc_info=True)              
                failed_log = {
                    "report_id": str(uuid.uuid4()),
                    "report_type": f"hubspot_{entity_type if 'entity_type' in locals() else 'unknown'}",
                    "requester_id": f"user:{clean_sender_email}",
                    "request_timestamp": datetime.now(pytz.timezone("Asia/Kolkata")).isoformat(),
                    "partial": True,
                    "failures": [{"error": str(report_error), "timestamp": datetime.now().isoformat()}]
                }
                logging.error(f"REPORT_FAILED: {json.dumps(failed_log)}")
                
                report_success = False
                ai_response = None
                report_filepath = None
           
            # Determine which response to use
            if report_success and ai_response:
                final_response = ai_response
                log_prefix = "SUCCESS Report"
            else:
                raise
                
           
            # Build and send email (keep your existing send logic)
            try:
                msg = MIMEMultipart()
                msg["From"] = f"HubSpot via lowtouch.ai <{HUBSPOT_FROM_ADDRESS}>"
                msg["To"] = sender_email
               
                subject = headers.get("Subject", "Your HubSpot Report")
                if not subject.lower().startswith("re:"):
                    subject = f"Re: {subject}"
                msg["Subject"] = subject
               
                original_message_id = headers.get("Message-ID", "")
                if original_message_id:
                    msg["In-Reply-To"] = original_message_id
                    references = headers.get("References", "")
                    msg["References"] = f"{references} {original_message_id}".strip() if references else original_message_id
               
                msg.attach(MIMEText(final_response, "html"))
               
                if report_success and report_filepath and os.path.exists(report_filepath):
                    with open(report_filepath, "rb") as f:
                        part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                        part.set_payload(f.read())
                    encoders.encode_base64(part)
                    part.add_header("Content-Disposition", f"attachment; filename={report_filename}")
                    msg.attach(part)
                    logging.info(f"✓ Attached report file: {report_filename}")
               
                raw = base64.urlsafe_b64encode(msg.as_string().encode()).decode()
                service.users().messages().send(userId="me", body={"raw": raw}).execute()                
                # Try to mark as read, but don't fail if it doesn't work
                try:
                    mark_message_as_read(service, email_id)
                except Exception as mark_error:
                    logging.warning(f"Could not mark email {email_id} as read: {mark_error}")             
                logging.info(f"✓ {log_prefix} email sent to {sender_email}")
           
            except Exception as send_error:
                logging.error(f"Failed to send email for {email_id}: {send_error}", exc_info=True)
                try:
                    mark_message_as_read(service, email_id)
                except:
                    pass
      
        except Exception as outer_error:
            logging.error(f"Unexpected error processing report email {email.get('id', 'unknown')}: {outer_error}", exc_info=True)
            try:
                mark_message_as_read(service, email.get("id", ""))
            except:
                pass
   
    logging.info(f"✓ Report DAG completed processing {len(report_emails)} emails")

def no_email_found(**kwargs):
    logging.info("No new emails or replies found to process.")

readme_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), 'hubspot_monitor_meeting_minutes.md')
readme_content = "HubSpot Meeting Minutes Mailbox Monitor DAG"
try:
    with open(readme_path, 'r') as file:
        readme_content = file.read()
except FileNotFoundError:
    pass

with DAG(
    "hubspot_monitor_mailbox",
    default_args=default_args,
    schedule=timedelta(minutes=1),
    catchup=False,
    doc_md=readme_content,
    tags=["hubspot", "monitor", "email", "mailbox"],
    on_success_callback=clear_retry_tracker_on_success,
    on_failure_callback=update_retry_tracker_on_failure
) as dag:

    fetch_emails_task = PythonOperator(
        task_id="fetch_unread_emails",
        python_callable=fetch_unread_emails,
    )

    branch_task = BranchPythonOperator(
        task_id="branch_task",
        python_callable=branch_function,
    )

    trigger_meeting_minutes_task = PythonOperator(
        task_id="trigger_meeting_minutes",
        python_callable=trigger_meeting_minutes,
    )

    trigger_continuation_task = PythonOperator(
        task_id="trigger_continuation_dag",
        python_callable=trigger_continuation_dag,
    )

    decide_and_search = PythonOperator(
        task_id='analyze_and_search_with_tools',
        python_callable=analyze_and_search_with_tools,
    )
    
    # === NEW TASK 2: Format response or trigger report ===
    generate_response = PythonOperator(
        task_id='generate_final_response_or_trigger_report',
        python_callable=generate_final_response_or_trigger_report,
        provide_context=True,
    )

    trigger_report_task = PythonOperator(
        task_id="trigger_report_dag",
        python_callable=trigger_report_dag,
    )

    trigger_task_completion_task = PythonOperator(
        task_id="trigger_task_completion",
        python_callable=trigger_task_completion_dag,
    )

    no_email_found_task = PythonOperator(
        task_id="no_email_found_task",
        python_callable=no_email_found,
    )

    fetch_emails_task >> branch_task >> [trigger_meeting_minutes_task, trigger_continuation_task, decide_and_search, trigger_report_task, trigger_task_completion_task, no_email_found_task]
    fetch_emails_task >> branch_task >> decide_and_search >> generate_response